"""
╔══════════════════════════════════════════════════════════════╗
║          CommerceShop Lead Database Dashboard                ║
║          Run:  streamlit run app.py                          ║
╚══════════════════════════════════════════════════════════════╝
"""

import json
import os
import io
from collections import OrderedDict
from datetime import datetime

import openpyxl
import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components

# ╔══════════════════════════════════════════════════════════════╗
# ║  FILE CONFIGURATION                                         ║
# ║  Files sit next to app.py (no subfolder needed).            ║
# ║  The big HubSpot file auto-downloads from Google Drive.     ║
# ╚══════════════════════════════════════════════════════════════╝

APP_DIR = os.path.dirname(os.path.abspath(__file__))

# CSV 1: Main HubSpot export (auto-downloaded from Google Drive)
HUBSPOT_GDRIVE_ID = "1iEJV-vbJuOxdBi_p_INBP8B43uOOCsAH"
HUBSPOT_CSV = os.path.join(APP_DIR, "all-contacts.csv")

# CSV 2: TCS email marketing (uploaded to GitHub)
TCS_EMAIL_MKT_CSV = os.path.join(APP_DIR, "Copy of TCS opener vs non opener - COMBINED LIST.csv")

# CSV 3: BinaryWorks email marketing (uploaded to GitHub)
BW_EMAIL_MKT_CSV = os.path.join(APP_DIR, "Copy of Drupal data cleaning - Sheet3.csv")

# CSV 4: Individual contribution (uploaded to GitHub)
CONTRIBUTION_XLSX = os.path.join(APP_DIR, "Copy of Over all DB .xlsx")

# Future: Add more email marketing CSVs here
# CB_EMAIL_MKT_CSV = os.path.join(APP_DIR, "conversionbox_email_marketing.csv")


def download_from_gdrive(file_id, destination):
    """Download a large file from Google Drive with virus scan confirmation."""
    if os.path.exists(destination):
        return  # Already downloaded

    # For large files, add confirm=t to bypass virus scan warning
    URL = f"https://drive.google.com/uc?export=download&id={file_id}&confirm=t"
    session = requests.Session()
    response = session.get(URL, stream=True)

    # Verify we got actual data, not HTML
    content_type = response.headers.get("content-type", "")
    if "text/html" in content_type:
        # Try alternative download URL
        URL2 = f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t"
        response = session.get(URL2, stream=True)

    with open(destination, "wb") as f:
        for chunk in response.iter_content(32768):
            if chunk:
                f.write(chunk)

    # Verify downloaded file is not HTML
    with open(destination, "r", errors="ignore") as f:
        first_line = f.readline(200)
    if first_line.strip().startswith("<!DOCTYPE") or first_line.strip().startswith("<html"):
        os.remove(destination)
        raise Exception("Google Drive download failed. Make sure the file sharing is set to 'Anyone with the link'.")

# ╔══════════════════════════════════════════════════════════════╗
# ║  BRAND TAG MAPPING                                          ║
# ║  Map TAG column values to brand names.                      ║
# ║  "Others" tags show as cards, not technology breakdown.     ║
# ╚══════════════════════════════════════════════════════════════╝

BRAND_TAGS = {
    "TCS": "tcs",
    "BinaryWorks": "drupal",
    "ConversionBox": "conversionbox",
}

# Tags that go under "Others" tab (cards with just leads + websites)
OTHERS_TAGS = ["Manufacturing", "ConversionBox 200+ Products", "Higher Education", "Finance"]

# TCS technologies to show individually (rest become "Other Technologies")
TCS_MAIN_TECHS = ["Shopify", "BigCommerce", "WooCommerce", "Magento", "Shopify Plus"]

# Drupal categories to show individually (rest become "Other CMS")
DRUPAL_MAIN_CATS = ["Drupal 7", "Drupal 8", "Drupal 9", "Drupal 10", "Drupal 11", "WordPress"]


# ══════════════════════════════════════════════════════════════
# STREAMLIT CONFIG
# ══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Lead Database Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
#MainMenu,header,footer{visibility:hidden}
.stApp{background-color:#050510}
.block-container{padding:0!important;max-width:100%!important}
iframe{border:none!important}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════

def safe_str(val):
    """Clean string value."""
    if pd.isna(val) or val is None:
        return ""
    return str(val).strip()


def load_hubspot(path):
    """Load main HubSpot CSV and build brand data."""
    if not os.path.exists(path):
        return None

    df = pd.read_csv(path, low_memory=False)

    # Normalize column names (handle slight variations)
    col_map = {}
    for c in df.columns:
        cl = c.strip().lower()
        if cl == "tag":
            col_map["tag"] = c
        elif cl == "email":
            col_map["email"] = c
        elif cl == "website url":
            col_map["website_url"] = c
        elif "e-commerce" in cl or "ecommerce" in cl:
            col_map["ecom_tech"] = c
        elif "drupal" in cl and "cms" in cl:
            col_map["drupal_cms"] = c
        elif "conversionbox" in cl and "competitor" in cl:
            col_map["cb_competitors"] = c

    if "email" not in col_map or "tag" not in col_map:
        st.error(f"HubSpot CSV must have 'Email' and 'TAG' columns. Found: {list(df.columns)}")
        return None

    result = {
        "tcs": {"rows": [], "totals": {"leads": 0, "websites": 0}, "emails": set()},
        "drupal": {"rows": [], "totals": {"leads": 0, "websites": 0}, "emails": set()},
        "conversionbox": {"rows": [], "totals": {"leads": 0, "websites": 0}, "emails": set()},
        "others": {},
        "overlap": {},
    }

    # Track which emails belong to which brands
    email_brands = {}

    for _, row in df.iterrows():
        tag = safe_str(row.get(col_map.get("tag", ""), ""))
        email = safe_str(row.get(col_map.get("email", ""), ""))
        website = safe_str(row.get(col_map.get("website_url", ""), ""))
        ecom = safe_str(row.get(col_map.get("ecom_tech", ""), ""))
        drupal = safe_str(row.get(col_map.get("drupal_cms", ""), ""))
        cb = safe_str(row.get(col_map.get("cb_competitors", ""), ""))

        if not email:
            continue

        # Track email -> brands for overlap
        if email not in email_brands:
            email_brands[email] = set()

        # TCS
        if tag in BRAND_TAGS and BRAND_TAGS[tag] == "tcs":
            result["tcs"]["emails"].add(email)
            email_brands[email].add("tcs")

        # BinaryWorks
        if tag in BRAND_TAGS and BRAND_TAGS[tag] == "drupal":
            result["drupal"]["emails"].add(email)
            email_brands[email].add("drupal")

        # ConversionBox
        if tag in BRAND_TAGS and BRAND_TAGS[tag] == "conversionbox":
            result["conversionbox"]["emails"].add(email)
            email_brands[email].add("conversionbox")

        # Others
        if tag in OTHERS_TAGS:
            if tag not in result["others"]:
                result["others"][tag] = {"emails": set(), "websites": set()}
            result["others"][tag]["emails"].add(email)
            if website:
                result["others"][tag]["websites"].add(website)

    # Build TCS technology breakdown
    tcs_df = df[df[col_map.get("tag", "")].apply(lambda x: safe_str(x)) == "TCS"]
    tcs_tech_counts = {}
    tcs_websites = {}

    for _, row in tcs_df.iterrows():
        tech = safe_str(row.get(col_map.get("ecom_tech", ""), ""))
        email = safe_str(row.get(col_map.get("email", ""), ""))
        website = safe_str(row.get(col_map.get("website_url", ""), ""))
        if not tech:
            tech = "Other Technologies"
        elif tech not in TCS_MAIN_TECHS:
            tech = "Other Technologies"

        if tech not in tcs_tech_counts:
            tcs_tech_counts[tech] = {"leads": 0, "websites": set(), "emails": set()}
        if email:
            tcs_tech_counts[tech]["leads"] += 1
            tcs_tech_counts[tech]["emails"].add(email)
        if website:
            tcs_tech_counts[tech]["websites"].add(website)

    tcs_rows = []
    for tech in TCS_MAIN_TECHS + ["Other Technologies"]:
        if tech in tcs_tech_counts:
            d = tcs_tech_counts[tech]
            tcs_rows.append({
                "label": tech,
                "leads": d["leads"],
                "websites": len(d["websites"]),
                "emails": list(d["emails"]),
            })
    result["tcs"]["rows"] = tcs_rows
    result["tcs"]["totals"] = {
        "leads": sum(r["leads"] for r in tcs_rows),
        "websites": len(set().union(*(set(tcs_tech_counts[t]["websites"]) for t in tcs_tech_counts))) if tcs_tech_counts else 0,
    }

    # Build Drupal breakdown
    drupal_df = df[df[col_map.get("tag", "")].apply(lambda x: safe_str(x)) == "BinaryWorks"]
    drupal_counts = {}

    for _, row in drupal_df.iterrows():
        cat = safe_str(row.get(col_map.get("drupal_cms", ""), ""))
        email = safe_str(row.get(col_map.get("email", ""), ""))
        website = safe_str(row.get(col_map.get("website_url", ""), ""))
        if not cat:
            cat = "Other CMS"
        elif cat not in DRUPAL_MAIN_CATS:
            cat = "Other CMS"

        if cat not in drupal_counts:
            drupal_counts[cat] = {"leads": 0, "websites": set(), "emails": set()}
        if email:
            drupal_counts[cat]["leads"] += 1
            drupal_counts[cat]["emails"].add(email)
        if website:
            drupal_counts[cat]["websites"].add(website)

    drupal_rows = []
    for cat in DRUPAL_MAIN_CATS + ["Other CMS"]:
        if cat in drupal_counts:
            d = drupal_counts[cat]
            drupal_rows.append({
                "label": cat,
                "leads": d["leads"],
                "websites": len(d["websites"]),
                "emails": list(d["emails"]),
            })
    result["drupal"]["rows"] = drupal_rows
    result["drupal"]["totals"] = {
        "leads": sum(r["leads"] for r in drupal_rows),
        "websites": len(set().union(*(set(drupal_counts[t]["websites"]) for t in drupal_counts))) if drupal_counts else 0,
    }

    # Build ConversionBox breakdown (show all values as-is)
    cb_df = df[df[col_map.get("tag", "")].apply(lambda x: safe_str(x)) == "ConversionBox"]
    cb_counts = {}

    for _, row in cb_df.iterrows():
        comp = safe_str(row.get(col_map.get("cb_competitors", ""), ""))
        email = safe_str(row.get(col_map.get("email", ""), ""))
        website = safe_str(row.get(col_map.get("website_url", ""), ""))
        if not comp:
            comp = "Unspecified"

        if comp not in cb_counts:
            cb_counts[comp] = {"leads": 0, "websites": set(), "emails": set()}
        if email:
            cb_counts[comp]["leads"] += 1
            cb_counts[comp]["emails"].add(email)
        if website:
            cb_counts[comp]["websites"].add(website)

    cb_rows = sorted(
        [{"label": k, "leads": v["leads"], "websites": len(v["websites"]), "emails": list(v["emails"])}
         for k, v in cb_counts.items()],
        key=lambda x: x["leads"], reverse=True
    )
    result["conversionbox"]["rows"] = cb_rows
    result["conversionbox"]["totals"] = {
        "leads": sum(r["leads"] for r in cb_rows),
        "websites": len(set().union(*(set(cb_counts[t]["websites"]) for t in cb_counts))) if cb_counts else 0,
    }

    # Build Others data
    others_data = []
    for tag in OTHERS_TAGS:
        if tag in result["others"]:
            d = result["others"][tag]
            others_data.append({
                "label": tag,
                "leads": len(d["emails"]),
                "websites": len(d["websites"]),
            })
    result["others"] = others_data

    # Build overlap data
    overlap = {"tcs": {}, "drupal": {}, "conversionbox": {}}
    for email, brands in email_brands.items():
        if len(brands) > 1:
            for b in brands:
                for other in brands:
                    if other != b:
                        if other not in overlap[b]:
                            overlap[b][other] = 0
                        overlap[b][other] += 1
    result["overlap"] = overlap

    # Clean emails from rows (convert sets for JSON)
    for brand in ["tcs", "drupal", "conversionbox"]:
        result[brand]["emails"] = list(result[brand]["emails"])
        for row in result[brand]["rows"]:
            row["emails"] = list(row.get("emails", []))

    return result


def load_email_marketing(csv_path, brand_emails_by_tech):
    """Load email marketing CSV and cross-reference with brand tech data."""
    if not os.path.exists(csv_path):
        return None

    df = pd.read_csv(csv_path, low_memory=False)

    # Find email and status columns
    email_col = None
    status_col = None
    for c in df.columns:
        cl = c.strip().lower()
        if cl == "email":
            email_col = c
        else:
            status_col = c  # The other column is the status

    if not email_col or not status_col:
        return None

    # Build lookup: email -> opener/non-opener
    mkt_data = {}
    for _, row in df.iterrows():
        email = safe_str(row.get(email_col, ""))
        status = safe_str(row.get(status_col, ""))
        if email:
            is_opener = "opener" in status.lower() and "non" not in status.lower()
            mkt_data[email] = "opener" if is_opener else "non-opener"

    # Cross-reference with tech rows
    result = {"total": 0, "openers": 0, "non_openers": 0, "by_tech": {}}
    mkt_emails_set = set(mkt_data.keys())

    for tech_row in brand_emails_by_tech:
        tech_label = tech_row["label"]
        tech_emails = set(tech_row.get("emails", []))
        matched = tech_emails & mkt_emails_set

        openers = sum(1 for e in matched if mkt_data.get(e) == "opener")
        non_openers = len(matched) - openers

        result["by_tech"][tech_label] = {
            "total": len(matched),
            "openers": openers,
            "non_openers": non_openers,
            "new_emails": len(tech_emails) - len(matched),
        }
        result["total"] += len(matched)
        result["openers"] += openers
        result["non_openers"] += non_openers

    all_brand_emails = set()
    for tr in brand_emails_by_tech:
        all_brand_emails.update(tr.get("emails", []))
    result["new_total"] = len(all_brand_emails) - len(all_brand_emails & mkt_emails_set)

    return result


def parse_contribution(xlsx_path):
    """Auto-detect format and parse individual contribution sheets."""
    if not os.path.exists(xlsx_path):
        return None

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    persons = {}

    MONTH_NAMES = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                   "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            all_rows.append(list(row))

        if not all_rows:
            continue

        # Auto-detect columns
        header = all_rows[0] if all_rows else []
        date_col = count_col = category_col = link_col = None
        month_col = None

        for i, h in enumerate(header):
            if h is None:
                continue
            hl = str(h).strip().lower()
            if hl in ("date", "date "):
                date_col = i
            elif hl in ("month",):
                month_col = i
            elif hl in ("count", "lead count", "lead_count", " count "):
                count_col = i
            elif "url" in hl or "link" in hl or "hubspot" in hl or "gmail" in hl:
                link_col = i
            elif hl in ("technology", "domain name", "domain", "category"):
                category_col = i

        # If no header matched, try positional detection
        if count_col is None:
            for i, h in enumerate(header):
                if h is None:
                    continue
                hl = str(h).strip().lower()
                if "count" in hl:
                    count_col = i
                    break

        # Parse data rows
        months = OrderedDict()
        current_month_key = None

        for row_idx, row in enumerate(all_rows[1:], start=2):
            # Handle Kishore-style format (month headers in data)
            cell0 = safe_str(row[0]) if row[0] else ""

            # Detect month header rows
            is_month_header = False
            for mn in MONTH_NAMES:
                if cell0.lower().startswith(mn) and ("month" in cell0.lower() or "data" in cell0.lower()):
                    current_month_key = mn.capitalize()
                    is_month_header = True
                    break

            if is_month_header:
                if current_month_key not in months:
                    months[current_month_key] = {"entries": []}
                continue

            # Skip total/summary rows
            if cell0.lower() in ("total count", "total", "summary", ""):
                # Try to catch total rows
                if cell0.lower() in ("total count", "total"):
                    continue
                if all(v is None for v in row):
                    continue

            # Try to extract count
            count_val = None
            if count_col is not None and count_col < len(row):
                try:
                    count_val = int(float(row[count_col]))
                except (ValueError, TypeError):
                    pass

            if count_val is None or count_val == 0:
                continue

            # Extract date/month
            date_str = ""
            month_key = current_month_key

            if date_col is not None and date_col < len(row) and row[date_col]:
                dv = row[date_col]
                if isinstance(dv, datetime):
                    month_key = dv.strftime("%b")
                    date_str = dv.strftime("%b %d")
                else:
                    date_str = safe_str(dv)
            elif month_col is not None and month_col < len(row) and row[month_col]:
                mv = safe_str(row[month_col])
                for mn in MONTH_NAMES:
                    if mv.lower().startswith(mn):
                        month_key = mn.capitalize()
                        break
                date_str = mv

            if not month_key:
                month_key = "Unknown"

            # Extract category
            category = ""
            if category_col is not None and category_col < len(row):
                category = safe_str(row[category_col])
            elif date_col is None and count_col is not None:
                # Kishore-style: category in first column
                category = cell0

            if not category:
                category = "General"

            # Normalize category
            if category.lower() in ("conversionbox", "conversionbox "):
                category = "ConversionBox"
            elif category.lower() in ("consultant", "consultant "):
                category = "Consultant"

            if month_key not in months:
                months[month_key] = {"entries": []}

            months[month_key]["entries"].append({
                "date": date_str or month_key,
                "category": category,
                "count": count_val,
            })

        # Sort months newest first
        sorted_months = sorted(
            months.keys(),
            key=lambda m: MONTH_NAMES.get(m.lower()[:3], 0),
            reverse=True,
        )

        person_months = []
        for mk in sorted_months:
            entries = months[mk]["entries"]
            if not entries:
                continue
            person_months.append({
                "month": f"{mk} 2026",
                "total": sum(e["count"] for e in entries),
                "entries": entries,
            })

        if person_months:
            # Derive person name from sheet name
            name = sheet_name.strip()
            # Try to extract a clean name
            for check in ["kishore", "ilakkiya", "illakkia", "dharanshri", "dharanishri"]:
                if check in name.lower():
                    name = check.capitalize()
                    if name == "Illakkia":
                        name = "Ilakkiya"
                    if name == "Dharanishri":
                        name = "Dharanshri"
                    break

            persons[name.lower().replace(" ", "_")] = {
                "name": name,
                "months": person_months,
            }

    wb.close()
    return persons if persons else None


def load_all_data():
    """Load all data sources."""
    data = {
        "brands": {
            "tcs": {"name": "TheCommerceShop", "colLabel": "Technology", "rows": [], "totals": {"leads": 0, "websites": 0}},
            "drupal": {"name": "BinaryWorks", "colLabel": "Category", "rows": [], "totals": {"leads": 0, "websites": 0}},
            "conversionbox": {"name": "ConversionBox", "colLabel": "Technology", "rows": [], "totals": {"leads": 0, "websites": 0}},
        },
        "others": [],
        "overlap": {"tcs": {}, "drupal": {}, "conversionbox": {}},
        "email_mkt": {"tcs": None, "drupal": None},
        "persons": {},
    }

    # Load HubSpot data (download from Google Drive if needed)
    if not os.path.exists(HUBSPOT_CSV):
        with st.spinner("Downloading HubSpot data from Google Drive (first time only)..."):
            download_from_gdrive(HUBSPOT_GDRIVE_ID, HUBSPOT_CSV)

    hub = load_hubspot(HUBSPOT_CSV)
    if hub:
        for brand in ["tcs", "drupal", "conversionbox"]:
            data["brands"][brand]["rows"] = hub[brand]["rows"]
            data["brands"][brand]["totals"] = hub[brand]["totals"]
        data["others"] = hub["others"]
        data["overlap"] = hub["overlap"]

        # Load email marketing
        tcs_mkt = load_email_marketing(TCS_EMAIL_MKT_CSV, hub["tcs"]["rows"])
        if tcs_mkt:
            data["email_mkt"]["tcs"] = tcs_mkt

        bw_mkt = load_email_marketing(BW_EMAIL_MKT_CSV, hub["drupal"]["rows"])
        if bw_mkt:
            data["email_mkt"]["drupal"] = bw_mkt

    # Load contribution data
    persons = parse_contribution(CONTRIBUTION_XLSX)
    if persons:
        data["persons"] = persons

    return data


# ══════════════════════════════════════════════════════════════
# HTML DASHBOARD TEMPLATE
# ══════════════════════════════════════════════════════════════

def build_dashboard_html(data):
    """Build the complete HTML dashboard."""

    # Clean data for JSON (remove email lists to reduce size)
    brands_clean = {}
    for bk, bv in data["brands"].items():
        brands_clean[bk] = {
            "name": bv["name"],
            "colLabel": bv["colLabel"],
            "rows": [{"label": r["label"], "leads": r["leads"], "websites": r["websites"]} for r in bv["rows"]],
            "totals": bv["totals"],
        }

    brands_json = json.dumps(brands_clean)
    others_json = json.dumps(data["others"])
    overlap_json = json.dumps(data["overlap"])
    email_mkt_json = json.dumps({
        k: {
            "total": v["total"],
            "openers": v["openers"],
            "non_openers": v["non_openers"],
            "new_total": v["new_total"],
            "by_tech": v["by_tech"],
        } if v else None
        for k, v in data["email_mkt"].items()
    })
    persons_json = json.dumps(data["persons"])

    return f"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
*{{margin:0;padding:0;box-sizing:border-box}}
:root{{
  --bg:#050510;--bg-card:#0c0c1d;--bg-elevated:#13132a;--border:#1c1c3a;
  --border-light:#181830;--text:#e8e8f0;--text-sec:#8888aa;--text-muted:#555570;
  --white:#fff;--accent:#4338ED;--accent2:#F97316;--accent-glow:rgba(67,56,237,0.3);
  --gradient:linear-gradient(135deg,#4338ED,#6366F1);
  --gradient-bar:linear-gradient(90deg,#4338ED,#818CF8);
  --stat-bg:linear-gradient(135deg,#0d0d1f,#111135);--stat-border:#252555;
}}
body{{font-family:'Inter',-apple-system,sans-serif;background:var(--bg);color:var(--text);line-height:1.5;
  background-image:radial-gradient(ellipse at 20% 0%,rgba(67,56,237,0.06) 0%,transparent 50%),
  radial-gradient(ellipse at 80% 100%,rgba(99,102,241,0.04) 0%,transparent 50%)}}
.db{{max-width:1180px;margin:0 auto;padding:36px 24px 56px}}
.hdr{{display:flex;align-items:center;justify-content:space-between;margin-bottom:32px;flex-wrap:wrap;gap:16px}}
.hdr h1{{font-size:26px;font-weight:900;color:var(--white)}}
.hdr h1 .hl{{color:var(--accent)}}
.hdr .sub{{font-size:13px;color:var(--text-sec);margin-top:3px}}
.tg{{display:inline-flex;background:var(--bg-card);border-radius:12px;padding:4px;border:1px solid var(--border)}}
.tb{{padding:9px 22px;border:none;background:transparent;font-family:inherit;font-size:13px;font-weight:600;color:var(--text-sec);border-radius:9px;cursor:pointer;transition:all .25s;white-space:nowrap}}
.tb:hover{{color:var(--white);background:var(--bg-elevated)}}
.tb.on{{background:var(--accent);color:var(--white);box-shadow:0 4px 16px var(--accent-glow)}}
.sr{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;margin-bottom:28px}}
.sc{{background:var(--stat-bg);border-radius:16px;padding:22px 24px;border:1px solid var(--stat-border);position:relative;overflow:hidden}}
.sc::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:var(--gradient)}}
.sc .lb{{font-size:10px;font-weight:700;color:var(--text-sec);text-transform:uppercase;letter-spacing:.7px;margin-bottom:8px}}
.sc .vl{{font-size:30px;font-weight:900;color:var(--white)}}
.sc .ic{{position:absolute;top:18px;right:20px;width:36px;height:36px;border-radius:10px;background:rgba(99,102,241,.15);border:1px solid rgba(99,102,241,.25);display:flex;align-items:center;justify-content:center;font-size:15px}}
.pn{{background:var(--bg-card);border-radius:18px;padding:28px;border:1px solid var(--border);margin-bottom:28px}}
.pt{{font-size:17px;font-weight:800;color:var(--white);margin-bottom:22px}}
.bc{{display:flex;flex-direction:column;gap:10px}}
.br{{display:grid;grid-template-columns:150px 1fr 80px 80px;align-items:center;gap:14px;padding:5px 0;border-radius:8px;transition:background .15s}}
.br.has-mkt{{grid-template-columns:150px 1fr 80px 80px 90px 90px}}
.br:hover{{background:rgba(255,255,255,.02)}}
.bl{{font-size:13px;font-weight:600;color:var(--text);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.bt{{height:30px;background:var(--bg-elevated);border-radius:8px;overflow:hidden}}
.bf{{height:100%;border-radius:8px;transition:width .7s cubic-bezier(.22,1,.36,1);min-width:6px;position:relative}}
.bf::after{{content:'';position:absolute;top:0;right:0;width:40px;height:100%;background:linear-gradient(90deg,transparent,rgba(255,255,255,.18));border-radius:0 8px 8px 0}}
.bv{{font-size:13px;font-weight:800;color:var(--white);text-align:right;font-variant-numeric:tabular-nums}}
.bv.sec{{color:var(--text-sec);font-weight:600}}
.bv.mkt{{color:var(--accent2);font-weight:700}}
.bv.new-em{{color:#34D399;font-weight:700}}
.bch{{font-size:10px;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:.5px;text-align:right}}
.bch.l{{text-align:left}}
.tr{{display:grid;grid-template-columns:150px 1fr 80px 80px;align-items:center;gap:14px;margin-top:16px;padding-top:16px;border-top:2px solid var(--accent)}}
.tr.has-mkt{{grid-template-columns:150px 1fr 80px 80px 90px 90px}}
.tr .bl{{font-weight:900;color:var(--accent);font-size:14px}}
.tr .bv{{font-weight:900;color:var(--accent)}}
.tr .bv.sec{{color:var(--accent2);font-weight:700}}
.overlap-tag{{display:inline-block;padding:2px 8px;border-radius:5px;font-size:9px;font-weight:700;background:rgba(249,115,22,.12);color:#F97316;border:1px solid rgba(249,115,22,.3);margin-left:6px}}
.mkt-toggle{{display:inline-flex;background:var(--bg);border-radius:8px;padding:2px;border:1px solid var(--border);margin-left:12px}}
.mkt-toggle button{{padding:4px 10px;border:none;background:transparent;font-family:inherit;font-size:10px;font-weight:600;color:var(--text-muted);border-radius:6px;cursor:pointer;transition:all .2s}}
.mkt-toggle button.on{{background:var(--accent2);color:var(--white)}}
.mkt-sec{{font-size:11px;font-weight:700;color:var(--accent2);margin:16px 0 10px;padding:8px 0;border-top:1px dashed var(--border);text-transform:uppercase;letter-spacing:.6px;display:flex;align-items:center;gap:8px}}

/* Others Cards */
.oc-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px}}
.oc{{background:var(--stat-bg);border-radius:16px;padding:24px;border:1px solid var(--stat-border);position:relative;overflow:hidden}}
.oc::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:var(--gradient)}}
.oc .oc-name{{font-size:14px;font-weight:800;color:var(--white);margin-bottom:12px}}
.oc .oc-row{{display:flex;justify-content:space-between;margin-bottom:6px}}
.oc .oc-label{{font-size:11px;color:var(--text-sec);font-weight:600}}
.oc .oc-val{{font-size:14px;font-weight:800;color:var(--white)}}

/* Contrib */
.cb-wrap{{display:flex;justify-content:center;margin:44px 0 28px}}
.cb-btn{{display:inline-flex;align-items:center;gap:10px;padding:14px 40px;background:var(--bg-card);border:1px solid var(--border);border-radius:14px;font-family:inherit;font-size:15px;font-weight:700;color:var(--white);cursor:pointer;transition:all .25s}}
.cb-btn:hover{{border-color:var(--accent);box-shadow:0 0 24px var(--accent-glow);background:var(--bg-elevated)}}
.cb-btn.on{{background:var(--accent);border-color:transparent;box-shadow:0 4px 20px var(--accent-glow)}}
.cb-btn .arr{{display:inline-block;transition:transform .3s;font-size:11px}}
.cb-btn.on .arr{{transform:rotate(180deg)}}
.cs{{display:none}}.cs.open{{display:block;animation:fi .35s ease}}
@keyframes fi{{from{{opacity:0;transform:translateY(-12px)}}to{{opacity:1;transform:translateY(0)}}}}
.ptw{{display:flex;justify-content:center;margin-bottom:28px}}
.gp{{background:var(--bg-card);border-radius:18px;padding:28px;border:1px solid var(--border);margin-bottom:24px}}
.gt{{font-size:15px;font-weight:800;color:var(--white);margin-bottom:24px}}
.ga{{position:relative;height:240px;display:flex;align-items:flex-end;padding-bottom:36px;padding-left:60px}}
.gy{{position:absolute;left:0;top:0;bottom:36px;width:55px;display:flex;flex-direction:column;justify-content:space-between;align-items:flex-end;padding-right:8px}}
.yl{{font-size:10px;color:var(--text-muted);font-variant-numeric:tabular-nums;font-weight:600}}
.gg{{position:absolute;left:60px;right:0;top:0;bottom:36px}}
.gl{{position:absolute;left:0;right:0;height:1px;background:var(--border-light);opacity:.5}}
.gbs{{display:flex;align-items:flex-end;gap:16px;flex:1;height:calc(100% - 36px);position:relative;z-index:1}}
.gbg{{flex:1;display:flex;flex-direction:column;align-items:center;height:100%;justify-content:flex-end;position:relative}}
.gb{{width:100%;max-width:64px;border-radius:10px 10px 4px 4px;position:relative;transition:all .5s cubic-bezier(.22,1,.36,1);cursor:pointer}}
.gb:hover{{filter:brightness(1.25);transform:scaleY(1.02);transform-origin:bottom}}
.gb .gv{{position:absolute;top:-24px;left:50%;transform:translateX(-50%);font-size:12px;font-weight:800;color:var(--white);white-space:nowrap}}
.gbl{{position:absolute;bottom:-30px;left:50%;transform:translateX(-50%);font-size:11px;font-weight:700;color:var(--text-sec);white-space:nowrap}}
.pp{{background:var(--bg-card);border-radius:18px;border:1px solid var(--border);margin-bottom:20px;overflow:hidden;display:none}}
.pp.vis{{display:block}}
.pph{{padding:18px 24px;border-bottom:1px solid var(--border);background:var(--bg-elevated)}}
.ppn{{font-size:17px;font-weight:800;color:var(--white)}}
.mr{{border-bottom:1px solid var(--border-light)}}.mr:last-child{{border-bottom:none}}
.mh{{display:flex;align-items:center;justify-content:space-between;padding:14px 24px;cursor:pointer;transition:background .15s;user-select:none}}
.mh:hover{{background:var(--bg-elevated)}}
.mh .ml{{display:flex;align-items:center;gap:10px}}
.mh .ma{{font-size:11px;color:var(--text-muted);transition:transform .2s;display:inline-block;width:16px;text-align:center}}
.mh .ma.op{{transform:rotate(90deg);color:var(--accent)}}
.mh .mm{{font-size:14px;font-weight:700;color:var(--white)}}
.mh .mc{{font-size:15px;font-weight:800;color:var(--white);font-variant-numeric:tabular-nums}}
.mh .me{{font-size:11px;color:var(--text-muted);margin-left:8px;font-weight:500}}
.md{{display:none;background:rgba(0,0,0,.25);border-top:1px solid var(--border-light)}}.md.op{{display:block}}
.dr{{display:grid;grid-template-columns:100px 1fr 100px;padding:11px 24px 11px 50px;gap:12px;border-bottom:1px solid rgba(255,255,255,.03);align-items:center}}
.dr:last-child{{border-bottom:none}}
.dd{{font-size:12px;color:var(--text-sec);font-weight:600;font-variant-numeric:tabular-nums}}
.dt{{display:inline-block;padding:3px 10px;border-radius:6px;font-size:10px;font-weight:700}}
.dc{{font-size:14px;font-weight:800;color:var(--white);text-align:right;font-variant-numeric:tabular-nums}}
.msw{{display:flex;justify-content:space-between;padding:11px 24px 11px 50px;background:var(--bg-elevated);border-top:1px solid var(--border)}}
.msw span{{font-size:13px;font-weight:800;color:var(--accent)}}
.no-data{{text-align:center;padding:48px 20px;color:var(--text-muted);font-size:14px}}
.ft{{text-align:center;padding:28px 0 0;font-size:11px;color:var(--text-muted)}}
@media(max-width:768px){{
  .db{{padding:16px 12px}}.hdr{{flex-direction:column;align-items:flex-start}}
  .sr{{grid-template-columns:1fr 1fr}}.br,.tr{{grid-template-columns:100px 1fr 60px 60px;gap:8px}}
  .br.has-mkt,.tr.has-mkt{{grid-template-columns:100px 1fr 60px 60px 60px 60px}}
  .pn,.gp{{padding:20px 16px}}.sc .vl{{font-size:24px}}
}}
@media(max-width:480px){{.br,.tr{{grid-template-columns:80px 1fr 50px}}.br .sec,.tr .sec,.bch.hm{{display:none}}}}
</style>
</head>
<body>
<div class="db" id="app"></div>
<script>
const THEMES={{
  tcs:{{accent:'#4338ED',accent2:'#F97316',glow:'rgba(67,56,237,0.3)',
    gradBar:'linear-gradient(90deg,#4338ED,#818CF8)',statBg:'linear-gradient(135deg,#0d0d1f,#111135)',statBdr:'#252555',chart:['#4338ED','#818CF8']}},
  drupal:{{accent:'#8B5CF6',accent2:'#F59E0B',glow:'rgba(139,92,246,0.3)',
    gradBar:'linear-gradient(90deg,#8B5CF6,#C4B5FD)',statBg:'linear-gradient(135deg,#0d0d1f,#1a1135)',statBdr:'#302555',chart:['#8B5CF6','#C4B5FD']}},
  conversionbox:{{accent:'#2563EB',accent2:'#10B981',glow:'rgba(37,99,235,0.3)',
    gradBar:'linear-gradient(90deg,#2563EB,#60A5FA)',statBg:'linear-gradient(135deg,#080e1f,#0c1a40)',statBdr:'#152550',chart:['#2563EB','#60A5FA']}},
  others:{{accent:'#6B7280',accent2:'#9CA3AF',glow:'rgba(107,114,128,0.3)',
    gradBar:'linear-gradient(90deg,#6B7280,#9CA3AF)',statBg:'linear-gradient(135deg,#0d0d15,#15151f)',statBdr:'#252530',chart:['#6B7280','#9CA3AF']}}
}};
const B={brands_json};
const OT={others_json};
const OV={overlap_json};
const EM={email_mkt_json};
const P={persons_json};
let aB='tcs',cO=false,aP=Object.keys(P)[0]||'',oM={{}},mktF='all';
function fmt(n){{return n.toLocaleString('en-IN')}}
function theme(t){{const s=document.documentElement.style,th=THEMES[t]||THEMES.others;
  s.setProperty('--accent',th.accent);s.setProperty('--accent2',th.accent2);
  s.setProperty('--accent-glow',th.glow);s.setProperty('--gradient-bar',th.gradBar);
  s.setProperty('--gradient','linear-gradient(135deg,'+th.accent+','+th.accent+'aa)');
  s.setProperty('--stat-bg',th.statBg);s.setProperty('--stat-border',th.statBdr)}}
function cc(c){{return{{'ConversionBox':'#818CF8','Drupal':'#60A5FA','TCS':'#6366F1','Consultant':'#C084FC','Freelancer':'#E879F9','ABM-Fintech':'#22D3EE','Tidio':'#34D399','Manufacturing List':'#FBBF24','TCS New Lead List':'#60A5FA','TCS New lead List':'#60A5FA','WordPress':'#38BDF8','Other CMS':'#FB923C','General':'#9CA3AF'}}[c]||'#9CA3AF'}}
function sB(k){{aB=k;mktF='all';render()}}
function tC(){{cO=!cO;render()}}
function sP(k){{aP=k;render()}}
function tM(p,m){{const k=p+'-'+m;oM[k]=!oM[k];render()}}
function sMF(f){{mktF=f;render()}}

function chart(p,th){{
  const ms=[...p.months].reverse(),mx=Math.max(...ms.map(m=>m.total));
  const st=4,sv=Math.ceil(mx/st/1000)*1000,cl=sv*st;
  let g='',y='';
  for(let i=0;i<=st;i++){{const pc=(i/st)*100,v=sv*i;g+=`<div class="gl" style="bottom:${{pc}}%"></div>`;y+=`<div class="yl">${{v>=1000?(v/1000).toFixed(0)+'K':v}}</div>`}}
  const[c1,c2]=th.chart;
  const bs=ms.map(m=>{{const h=cl>0?(m.total/cl)*100:0;return`<div class="gbg"><div class="gb" style="height:${{Math.max(h,3)}}%;background:linear-gradient(180deg,${{c2}},${{c1}});box-shadow:0 0 16px ${{th.glow}}"><div class="gv">${{fmt(m.total)}}</div><div class="gbl">${{m.month.split(' ')[0]}}</div></div></div>`}}).join('');
  return`<div class="gp"><div class="gt">${{p.name}}'s Monthly Contribution</div><div class="ga"><div class="gy">${{y}}</div><div class="gg">${{g}}</div><div class="gbs">${{bs}}</div></div></div>`;
}}

function render(){{
  const b=B[aB],th=THEMES[aB]||THEMES.others;theme(aB);
  const isOthers=aB==='others';

  // Overlap text for this brand
  const ovl=OV[aB]||{{}};
  let ovlHTML='';
  for(const[ob,cnt] of Object.entries(ovl)){{
    const names={{tcs:'TCS',drupal:'BinaryWorks',conversionbox:'ConversionBox'}};
    ovlHTML+=`<span class="overlap-tag">${{fmt(cnt)}} also in ${{names[ob]||ob}}</span>`;
  }}

  // Email marketing data
  const mkt=EM[aB]||null;
  const hasMkt=!!mkt;
  const mktCls=hasMkt?'has-mkt':'';

  let mainHTML='';
  if(isOthers){{
    // Others view - cards
    const stats=`<div class="sr">
      <div class="sc"><div class="lb">Total Tags</div><div class="vl">${{OT.length}}</div><div class="ic">🏷️</div></div>
      <div class="sc"><div class="lb">Total Leads</div><div class="vl">${{fmt(OT.reduce((s,o)=>s+o.leads,0))}}</div><div class="ic">📊</div></div>
      <div class="sc"><div class="lb">Total Websites</div><div class="vl">${{fmt(OT.reduce((s,o)=>s+o.websites,0))}}</div><div class="ic">🌐</div></div>
    </div>`;
    const cards=OT.length?`<div class="oc-grid">${{OT.map(o=>`<div class="oc"><div class="oc-name">${{o.label}}</div><div class="oc-row"><span class="oc-label">Leads</span><span class="oc-val">${{fmt(o.leads)}}</span></div><div class="oc-row"><span class="oc-label">Websites</span><span class="oc-val">${{fmt(o.websites)}}</span></div></div>`).join('')}}</div>`:`<div class="no-data">No data files found. Place your HubSpot CSV at:<br><code>${HUBSPOT_CSV.replace(os.sep, '/')}</code></div>`;
    mainHTML=stats+cards;
  }} else {{
    if(!b || !b.rows.length){{
      mainHTML=`<div class="sr"><div class="sc"><div class="lb">Total Leads</div><div class="vl">0</div></div></div><div class="no-data">No data found for this brand. Place your HubSpot CSV at:<br><code>${HUBSPOT_CSV.replace(os.sep, '/')}</code></div>`;
    }} else {{
      const maxL=Math.max(...b.rows.map(r=>r.leads));
      // Stats
      let statsHTML=`<div class="sr">
        <div class="sc"><div class="lb">Total Leads</div><div class="vl">${{fmt(b.totals.leads)}}</div><div class="ic">📊</div></div>
        <div class="sc"><div class="lb">Unique Websites</div><div class="vl">${{fmt(b.totals.websites)}}</div><div class="ic">🌐</div></div>
        <div class="sc"><div class="lb">Avg Leads / Website</div><div class="vl">${{(b.totals.leads/(b.totals.websites||1)).toFixed(1)}}</div><div class="ic">⚡</div></div>`;
      if(hasMkt){{
        statsHTML+=`<div class="sc"><div class="lb">Email Mkt Using</div><div class="vl">${{fmt(mkt.total)}}</div><div class="ic">📧</div></div>
        <div class="sc"><div class="lb">New Emails</div><div class="vl">${{fmt(mkt.new_total)}}</div><div class="ic">✨</div></div>`;
      }}
      statsHTML+=`</div>`;
      if(ovlHTML) statsHTML+=`<div style="margin:-16px 0 20px;padding:0 4px">${{ovlHTML}}</div>`;

      // Mkt filter toggle
      let mktToggle='';
      if(hasMkt){{
        mktToggle=`<div class="mkt-toggle"><button class="${{mktF==='all'?'on':''}}" onclick="sMF('all')">All</button><button class="${{mktF==='opener'?'on':''}}" onclick="sMF('opener')">Openers</button><button class="${{mktF==='non-opener'?'on':''}}" onclick="sMF('non-opener')">Non-Openers</button></div>`;
      }}

      // Bar chart headers
      let headHTML=`<div class="br ${{mktCls}}" style="margin-bottom:4px"><div class="bch l">${{b.colLabel}}</div><div></div><div class="bch">Leads</div><div class="bch hm">Websites</div>`;
      if(hasMkt) headHTML+=`<div class="bch">Mkt Using</div><div class="bch">New</div>`;
      headHTML+=`</div>`;

      // Bar rows
      let rowsHTML=b.rows.map(r=>{{
        let mktVal='',newVal='';
        if(hasMkt && mkt.by_tech[r.label]){{
          const mt=mkt.by_tech[r.label];
          if(mktF==='all') mktVal=fmt(mt.total);
          else if(mktF==='opener') mktVal=fmt(mt.openers);
          else mktVal=fmt(mt.non_openers);
          newVal=fmt(mt.new_emails);
        }}
        return`<div class="br ${{mktCls}}"><div class="bl" title="${{r.label}}">${{r.label}}</div><div class="bt"><div class="bf" style="width:${{(r.leads/maxL*100).toFixed(1)}}%;background:var(--gradient-bar)"></div></div><div class="bv">${{fmt(r.leads)}}</div><div class="bv sec">${{fmt(r.websites)}}</div>${{hasMkt?`<div class="bv mkt">${{mktVal}}</div><div class="bv new-em">${{newVal}}</div>`:''}}</div>`;
      }}).join('');

      // Totals
      let totMkt='',totNew='';
      if(hasMkt){{
        if(mktF==='all') totMkt=fmt(mkt.total);
        else if(mktF==='opener') totMkt=fmt(mkt.openers);
        else totMkt=fmt(mkt.non_openers);
        totNew=fmt(mkt.new_total);
      }}
      let totHTML=`<div class="tr ${{mktCls}}"><div class="bl">TOTAL</div><div></div><div class="bv">${{fmt(b.totals.leads)}}</div><div class="bv sec">${{fmt(b.totals.websites)}}</div>${{hasMkt?`<div class="bv mkt">${{totMkt}}</div><div class="bv new-em">${{totNew}}</div>`:''}}</div>`;

      mainHTML=statsHTML+`<div class="pn"><div style="display:flex;align-items:center;flex-wrap:wrap;margin-bottom:22px"><div class="pt" style="margin:0">${{b.colLabel}} Breakdown</div>${{mktToggle}}</div><div class="bc">${{headHTML}}${{rowsHTML}}${{totHTML}}</div></div>`;
    }}
  }}

  // Person panels
  let pnlHTML='';
  if(Object.keys(P).length){{
    for(const[pk,p] of Object.entries(P)){{
      let mHTML='';
      p.months.forEach((m,mi)=>{{const op=oM[pk+'-'+mi];
        mHTML+=`<div class="mr"><div class="mh" onclick="tM('${{pk}}',${{mi}})"><div class="ml"><span class="ma ${{op?'op':''}}">&#9654;</span><span class="mm">${{m.month}}</span><span class="me">${{m.entries.length}} ${{m.entries.length===1?'entry':'entries'}}</span></div><span class="mc">${{fmt(m.total)}}</span></div><div class="md ${{op?'op':''}}">
        ${{m.entries.map(e=>`<div class="dr"><div class="dd">${{e.date}}</div><div><span class="dt" style="background:${{cc(e.category)}}18;color:${{cc(e.category)}};border:1px solid ${{cc(e.category)}}40">${{e.category}}</span></div><div class="dc">${{fmt(e.count)}}</div></div>`).join('')}}
        <div class="msw"><span>Subtotal</span><span>${{fmt(m.total)}}</span></div></div></div>`;
      }});
      pnlHTML+=`<div class="pp ${{aP===pk?'vis':''}}"><div class="pph"><span class="ppn">${{p.name}}</span></div>${{mHTML}}</div>`;
    }}
  }}

  const personKeys=Object.keys(P);
  const personBtns=personKeys.map(k=>`<button class="tb ${{aP===k?'on':''}}" onclick="sP('${{k}}')">${{P[k].name}}</button>`).join('');
  const hasPersons=personKeys.length>0;

  document.getElementById('app').innerHTML=`
    <div class="hdr"><div><h1>${{isOthers?'Others':B[aB]?.name||''}} <span class="hl">Lead Database</span></h1><div class="sub">Live data from HubSpot export</div></div>
    <div class="tg"><button class="tb ${{aB==='tcs'?'on':''}}" onclick="sB('tcs')">TCS</button><button class="tb ${{aB==='drupal'?'on':''}}" onclick="sB('drupal')">BinaryWorks</button><button class="tb ${{aB==='conversionbox'?'on':''}}" onclick="sB('conversionbox')">ConversionBox</button><button class="tb ${{aB==='others'?'on':''}}" onclick="sB('others')">Others</button></div></div>
    ${{mainHTML}}
    ${{hasPersons?`<div class="cb-wrap"><button class="cb-btn ${{cO?'on':''}}" onclick="tC()"><span class="arr">&#9660;</span>Individual Contribution</button></div>
    <div class="cs ${{cO?'open':''}}"><div class="ptw"><div class="tg">${{personBtns}}</div></div>
    ${{aP&&P[aP]?chart(P[aP],th):''}}${{pnlHTML}}</div>`:''}}
    <div class="ft">CommerceShop Lead Database Dashboard &middot; Data loaded live</div>`;
}}
render();
</script>
</body>
</html>
"""


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

data = load_all_data()
html = build_dashboard_html(data)
components.html(html, height=1800, scrolling=True)
