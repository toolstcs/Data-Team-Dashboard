"""
CommerceShop Lead Database Dashboard
Run: streamlit run app.py
"""

import json, os, io
from collections import OrderedDict
from datetime import datetime
import openpyxl, pandas as pd, requests
import streamlit as st
import streamlit.components.v1 as components

# ════════════════════════════════════════
# FILE CONFIG - update paths here
# ════════════════════════════════════════
APP_DIR = os.path.dirname(os.path.abspath(__file__))

# ── HubSpot Service Key (stored in Streamlit Secrets, not in code) ──
# Go to Streamlit > Manage app > Settings > Secrets and add:
# HUBSPOT_API_KEY = "your-key-here"
HUBSPOT_SERVICE_KEY = ""  # Will be loaded from Streamlit Secrets below

# HubSpot internal property names
HS_PROPS = {
    "email": "email",
    "website": "website",
    "tag": "tag",
    "ecom": "e_commerce_technologies",
    "drupal": "drupal_partners__cms_",
    "cb": "conversionbox_competitors",
    # Product Count is auto-resolved from its label at runtime (see
    # resolve_product_count_prop). This value is only the fallback guess used if
    # the live property list cannot be fetched.
    "products": "product_count",
}

# Human label of the Product Count column, used to auto-resolve its real
# internal name against the live HubSpot property list. HubSpot internal names
# cannot contain spaces, so the label and the internal name differ.
PRODUCT_COUNT_LABEL = "Product Count"

# Bands for the ConversionBox "200+ Products" section. Contacts with Product
# Count below 200 are dropped entirely. Bands are inclusive on both ends and do
# not overlap, so every kept contact lands in exactly one band. hi = None means
# open-ended (the top band).
PRODUCT_COUNT_MIN = 200
PRODUCT_BANDS = [
    {"label": "200 - 999", "lo": 200, "hi": 999},
    {"label": "1K - 9,999", "lo": 1000, "hi": 9999},
    {"label": "10K - 99,999", "lo": 10000, "hi": 99999},
    {"label": "100K+", "lo": 100000, "hi": None},
]

HUBSPOT_GDRIVE_ID = "1iEJV-vbJuOxdBi_p_INBP8B43uOOCsAH"
HUBSPOT_CSV = os.path.join(APP_DIR, "all-contacts.csv")
TCS_EMAIL_MKT_CSV = os.path.join(APP_DIR, "Copy of TCS opener vs non opener - COMBINED LIST.csv")
BW_EMAIL_MKT_CSV = os.path.join(APP_DIR, "Copy of Drupal data cleaning - Sheet3.csv")

# Extra email list (all brands, downloaded from Google Drive)
EXTRA_EMAIL_GDRIVE_ID = "1UE_rb7o5ODjSIpHJFSzqBNJHV37-Xht-"
EXTRA_EMAIL_CSV = os.path.join(APP_DIR, "tcs-email-using-combined.csv")
# Individual contribution: read directly from Google Sheets (no download needed)
GSHEET_PUB_KEY = "2PACX-1vRnDbY4gZeBoX_QspGXBhFVwy9f2kSuJ64XEvpgaFmcWtZqR7F8Dh6IeMr4q8khu-1Add_fyu93_hg0"
GSHEET_TABS = {
    "kishore": {"gid": 56889493, "name": "Kishore"},
    "ilakkiya": {"gid": 389171671, "name": "Ilakkiya"},
    "dharanshri": {"gid": 422158322, "name": "Dharanshri"},
    "sales": {"gid": 231515284, "name": "Sales"},
}

# Brand TAG mapping (all lowercase keys for case-insensitive matching)
BRAND_TAGS = {"tcs": "tcs", "binaryworks": "drupal", "drupal": "drupal",
              "conversionbox": "conversionbox", "conversionbox compitetor": "conversionbox"}

# "conversionbox 200+ products" was removed from Others: it is now a section
# inside the ConversionBox tab, driven by the Product Count column instead of
# this tag.
OTHERS_TAGS = ["manufacturing", "higher education", "fin tech"]
OTHERS_DISPLAY = {"manufacturing": "Manufacturing",
                  "higher education": "Higher Education", "fin tech": "Fin Tech"}

TCS_MAIN_TECHS = ["Shopify", "BigCommerce", "WooCommerce", "Magento", "Shopify Plus"]
DRUPAL_MAIN_CATS = ["Drupal 7", "Drupal 8", "Drupal 9", "Drupal 10", "Drupal 11", "WordPress"]

VALID_MONTHS = {"Jan", "Feb", "Mar", "Apr", "May", "Jun"}
MONTH_ORDER = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}

# ════════════════════════════════════════
# STREAMLIT CONFIG
# ════════════════════════════════════════
st.set_page_config(page_title="Lead Database Dashboard", page_icon="📊", layout="wide", initial_sidebar_state="collapsed")
st.markdown("""<style>#MainMenu,header,footer{visibility:hidden}.stApp{background-color:#f8f9fa}
.block-container{padding:0!important;max-width:100%!important}iframe{border:none!important}</style>""", unsafe_allow_html=True)

# Load HubSpot key from Streamlit Secrets
try:
    HUBSPOT_SERVICE_KEY = st.secrets["HUBSPOT_API_KEY"]
except:
    HUBSPOT_SERVICE_KEY = ""


# ════════════════════════════════════════
# GOOGLE DRIVE DOWNLOAD
# ════════════════════════════════════════
def download_gdrive(file_id, dest):
    if os.path.exists(dest):
        try:
            with open(dest, "r", errors="ignore") as f:
                line = f.readline(200)
            if not line.strip().startswith("<!") and not line.strip().startswith("<html") and os.path.getsize(dest) > 5000:
                return
        except:
            pass
        os.remove(dest)

    url = f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t"
    r = requests.get(url, stream=True)
    with open(dest, "wb") as f:
        for chunk in r.iter_content(32768):
            if chunk: f.write(chunk)

    with open(dest, "r", errors="ignore") as f:
        line = f.readline(200)
    if line.strip().startswith("<!") or line.strip().startswith("<html"):
        os.remove(dest)
        st.error("Google Drive download failed. Check the sharing link.")
        return


# ════════════════════════════════════════
# HUBSPOT API (Quick Count via Search)
# ════════════════════════════════════════
def hs_search_count(filters):
    """Search HubSpot contacts and return total count."""
    url = "https://api.hubapi.com/crm/v3/objects/contacts/search"
    headers = {"Authorization": f"Bearer {HUBSPOT_SERVICE_KEY}", "Content-Type": "application/json"}
    body = {"filterGroups": [{"filters": filters}], "limit": 1}
    try:
        r = requests.post(url, headers=headers, json=body, timeout=15)
        if r.status_code == 200:
            return r.json().get("total", 0)
        else:
            st.warning(f"HubSpot API error: {r.status_code}")
            return 0
    except Exception as e:
        st.warning(f"HubSpot connection error: {e}")
        return 0


def hs_prop_filter(prop, operator, value=None):
    """Build a HubSpot filter dict."""
    f = {"propertyName": prop, "operator": operator}
    if value is not None:
        f["value"] = value
    return f


def _norm_label(s):
    """Lowercase and strip everything that is not a letter or digit."""
    import re
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


@st.cache_data(ttl=86400, show_spinner=False)
def hs_list_contact_properties():
    """
    All contact properties as {internal_name: label}. Cached for a day, since
    the schema barely changes. Empty dict on any failure, so callers fall back
    to the guessed name rather than crash.
    """
    if not HUBSPOT_SERVICE_KEY:
        return {}
    url = "https://api.hubapi.com/crm/v3/properties/contacts"
    headers = {"Authorization": f"Bearer {HUBSPOT_SERVICE_KEY}"}
    try:
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code != 200:
            return {}
        return {p["name"]: p.get("label", p["name"]) for p in r.json().get("results", [])}
    except Exception:
        return {}


def resolve_product_count_prop():
    """
    Find the real internal name of the Product Count column.

    Order: trust the configured guess if it genuinely exists in the portal,
    otherwise match on the exact normalised label ("Product Count"), never a
    substring. Returns (internal_name_or_None, note_for_display).
    """
    props = hs_list_contact_properties()
    guess = HS_PROPS.get("products")

    if not props:
        # No live schema (no key, or the call failed). Use the guess and say so.
        return guess, f"assumed `{guess}` (could not read live properties)"

    if guess in props:
        return guess, f"`{guess}`"

    target = _norm_label(PRODUCT_COUNT_LABEL)
    matches = [name for name, label in props.items() if _norm_label(label) == target]
    if matches:
        # Prefer a non hs_ custom property if several share the label.
        matches.sort(key=lambda n: (n.startswith("hs_"), n))
        return matches[0], f"`{matches[0]}` (matched by label '{PRODUCT_COUNT_LABEL}')"

    return None, (
        f"NOT FOUND. No property is named `{guess}` or labelled "
        f"'{PRODUCT_COUNT_LABEL}'. The 200+ Products section will be empty."
    )


def hs_product_band_count(prop, lo, hi):
    """
    Count contacts whose Product Count falls in [lo, hi].

    HubSpot number properties support GTE/LTE. hi=None means open-ended (>= lo).
    Both filters sit in one group, so they are AND-ed.
    """
    filters = [hs_prop_filter(prop, "GTE", str(lo))]
    if hi is not None:
        filters.append(hs_prop_filter(prop, "LTE", str(hi)))
    return hs_search_count(filters)


def fetch_product_bands_live():
    """
    Fetch ONLY the ConversionBox "200+ Products" bands, live from HubSpot.

    This is the Option A path: everything else on the dashboard comes from the
    Google Drive CSV, but Product Count is read live from the API on every load,
    because the CSV export does not carry that column. Four API calls, one per
    band. Returns the same {rows, totals, note} shape the CSV path would have
    produced, so build_html does not care where it came from.
    """
    import time
    out = {"rows": [], "totals": {"leads": 0, "websites": 0}, "note": ""}

    if not HUBSPOT_SERVICE_KEY:
        out["note"] = "no HubSpot key set, 200+ Products cannot load"
        return out

    pc, note = resolve_product_count_prop()
    out["note"] = note
    if not pc:
        return out

    total = 0
    for band in PRODUCT_BANDS:
        c = hs_product_band_count(pc, band["lo"], band["hi"])
        # Keep every band, even zero, so the section shows the full ladder.
        out["rows"].append({"label": band["label"], "leads": c, "websites": 0, "emails": []})
        total += c
        time.sleep(0.3)
    out["totals"]["leads"] = total
    return out


@st.cache_data(ttl=1800, show_spinner=False)
def cached_product_bands():
    """
    Live product bands, cached for 30 minutes.

    Streamlit reruns the whole script on every widget interaction. Without this
    cache, opening a month row or switching a brand tab would refire the four
    Product Count API calls each time. The cache key is the function itself
    (no args), so all sessions share one 30-minute result. The Refresh button,
    which clears caches, still forces a fresh pull.
    """
    return fetch_product_bands_live()


def fetch_hubspot_counts():
    """Fetch live counts from HubSpot Search API."""
    import time
    result = {"tcs": {"rows": [], "totals": {"leads": 0, "websites": 0}},
              "drupal": {"rows": [], "totals": {"leads": 0, "websites": 0}},
              "conversionbox": {"rows": [], "totals": {"leads": 0, "websites": 0}},
              # products: the ConversionBox "200+ Products" section, banded by
              # Product Count. Kept separate from the competitors rows above.
              "products": {"rows": [], "totals": {"leads": 0, "websites": 0}, "note": ""},
              "others": [], "overlap": {"tcs": {}, "drupal": {}, "conversionbox": {}}}

    ep = HS_PROPS["ecom"]
    dp = HS_PROPS["drupal"]
    cp = HS_PROPS["cb"]
    tp = HS_PROPS["tag"]
    pc, pc_note = resolve_product_count_prop()

    # ── TCS: e-Commerce Technologies ──
    tcs_techs = {
        "Shopify": ["Shopify"],
        "BigCommerce": ["BigCommerce", "Big Commerce"],
        "WooCommerce": ["WooCommerce", "Woo Commerce"],
        "Magento": ["Magento"],
        "Shopify Plus": ["Shopify Plus"],
    }

    tcs_total = hs_search_count([hs_prop_filter(ep, "HAS_PROPERTY")])
    time.sleep(0.3)
    tcs_named = 0
    for tech, values in tcs_techs.items():
        count = 0
        for val in values:
            c = hs_search_count([hs_prop_filter(ep, "CONTAINS_TOKEN", val)])
            count += c
            time.sleep(0.3)
        # Special: Shopify should not include Shopify Plus
        if tech == "Shopify":
            plus_count = 0
            for val in tcs_techs.get("Shopify Plus", []):
                plus_count += hs_search_count([hs_prop_filter(ep, "CONTAINS_TOKEN", val)])
                time.sleep(0.3)
            count = max(0, count - plus_count)
        result["tcs"]["rows"].append({"label": tech, "leads": count, "websites": 0, "emails": []})
        tcs_named += count

    other_tcs = max(0, tcs_total - tcs_named)
    result["tcs"]["rows"].append({"label": "Other Technologies", "leads": other_tcs, "websites": 0, "emails": []})
    result["tcs"]["totals"]["leads"] = tcs_total

    # ── BinaryWorks: Drupal Partners (CMS) ──
    drupal_total = hs_search_count([hs_prop_filter(dp, "HAS_PROPERTY")])
    time.sleep(0.3)
    drupal_cats = {}
    for ver in ["7", "8", "9", "10", "11"]:
        c = hs_search_count([hs_prop_filter(dp, "CONTAINS_TOKEN", f"Drupal {ver}")])
        if c > 0: drupal_cats[f"Drupal {ver}"] = c
        time.sleep(0.3)

    # Drupal generic
    drupal_generic = hs_search_count([hs_prop_filter(dp, "EQ", "Drupal")])
    if drupal_generic > 0: drupal_cats["Drupal (Generic)"] = drupal_generic
    time.sleep(0.3)

    # WordPress (all versions)
    wp_count = hs_search_count([hs_prop_filter(dp, "CONTAINS_TOKEN", "WordPress")])
    wp_count2 = hs_search_count([hs_prop_filter(dp, "CONTAINS_TOKEN", "Wordpress")])
    wp_total = max(wp_count, wp_count2)
    if wp_total > 0: drupal_cats["WordPress"] = wp_total
    time.sleep(0.3)

    named_drupal = sum(drupal_cats.values())
    other_cms = max(0, drupal_total - named_drupal)
    if other_cms > 0: drupal_cats["Other CMS"] = other_cms

    for cat in ["Drupal 7", "Drupal 8", "Drupal 9", "Drupal 10", "Drupal 11", "Drupal (Generic)", "WordPress", "Other CMS"]:
        if cat in drupal_cats:
            result["drupal"]["rows"].append({"label": cat, "leads": drupal_cats[cat], "websites": 0, "emails": []})
    result["drupal"]["totals"]["leads"] = drupal_total

    # ── ConversionBox ──
    cb_values = ["Doofinder", "SearchSpring", "Klevu", "Swiftype", "LucidWorks",
                 "Algolia", "Fast Simon", "Searchanise", "Boost Commerce",
                 "Hawk search", "Elastic Suite", "Zopim"]
    cb_total = hs_search_count([hs_prop_filter(cp, "HAS_PROPERTY")])
    time.sleep(0.3)
    for val in cb_values:
        c = hs_search_count([hs_prop_filter(cp, "EQ", val)])
        if c > 0:
            result["conversionbox"]["rows"].append({"label": val, "leads": c, "websites": 0, "emails": []})
        time.sleep(0.3)
    result["conversionbox"]["totals"]["leads"] = cb_total

    # ── ConversionBox: 200+ Products (banded by Product Count) ──
    # Everything below PRODUCT_COUNT_MIN is excluded: the section total is the
    # sum of the bands, not a HAS_PROPERTY count, so contacts under 200 never
    # enter the numbers.
    result["products"]["note"] = pc_note
    if pc:
        prod_total = 0
        for band in PRODUCT_BANDS:
            c = hs_product_band_count(pc, band["lo"], band["hi"])
            result["products"]["rows"].append(
                {"label": band["label"], "leads": c, "websites": 0, "emails": []})
            prod_total += c
            time.sleep(0.3)
        result["products"]["totals"]["leads"] = prod_total

    # ── Others (TAG based) ──
    for tag_lower, display in OTHERS_DISPLAY.items():
        c = hs_search_count([hs_prop_filter(tp, "EQ", tag_lower)])
        if c > 0:
            result["others"].append({"label": display, "leads": c, "websites": 0})
        time.sleep(0.3)

    return result


# ════════════════════════════════════════
# FIND COLUMNS (flexible matching)
# ════════════════════════════════════════
def find_columns(df):
    m = {}
    for c in df.columns:
        cl = c.strip().lower()
        if cl == "tag": m["tag"] = c
        elif cl == "email": m["email"] = c
        elif cl == "website url": m["web"] = c
        elif cl == "e-commerce technologies": m["ecom"] = c
        elif cl == "drupal partners (cms)": m["drupal"] = c
        elif cl == "conversionbox competitors": m["cb"] = c
        elif cl == "product count": m["products"] = c
    # Fallback: flexible matching if exact names didn't work
    if "ecom" not in m:
        for c in df.columns:
            cl = c.strip().lower()
            if ("e-commerce" in cl or "ecommerce" in cl) and cl not in [v.strip().lower() for v in m.values()]:
                m["ecom"] = c; break
    if "drupal" not in m:
        for c in df.columns:
            cl = c.strip().lower()
            if cl.startswith("drupal") and ("cms" in cl or "partner" in cl):
                m["drupal"] = c; break
    if "cb" not in m:
        for c in df.columns:
            cl = c.strip().lower()
            if cl.startswith("conversionbox") and "compet" in cl:
                m["cb"] = c; break
    if "products" not in m:
        for c in df.columns:
            cl = c.strip().lower()
            if "product" in cl and "count" in cl:
                m["products"] = c; break
    return m


# ════════════════════════════════════════
# LOAD HUBSPOT DATA
# ════════════════════════════════════════
def load_hubspot():
    if not os.path.exists(HUBSPOT_CSV):
        return None

    df = pd.read_csv(HUBSPOT_CSV, low_memory=False, on_bad_lines="skip", encoding="utf-8", encoding_errors="replace")
    cols = find_columns(df)

    if "email" not in cols or "tag" not in cols:
        st.error(f"CSV needs Email and TAG columns. Found: {list(df.columns)[:10]}")
        return None

    tc, ec, wc = cols["tag"], cols["email"], cols.get("web","")
    ecom_c, drupal_c, cb_c = cols.get("ecom",""), cols.get("drupal",""), cols.get("cb","")
    prod_c = cols.get("products","")

    # Clean
    for c in [tc, ec]:
        df[c] = df[c].fillna("").astype(str).str.strip()
    if wc: df[wc] = df[wc].fillna("").astype(str).str.strip()
    if ecom_c: df[ecom_c] = df[ecom_c].fillna("").astype(str).str.strip()
    if drupal_c: df[drupal_c] = df[drupal_c].fillna("").astype(str).str.strip()
    if cb_c: df[cb_c] = df[cb_c].fillna("").astype(str).str.strip()

    df = df[df[ec] != ""]
    df["_tag_lower"] = df[tc].str.lower().str.strip()

    # Brand map
    df["_brand"] = df["_tag_lower"].map(BRAND_TAGS).fillna("")

    # ── TCS: rows where e-Commerce Technologies has a value ──
    tcs_df = df[df[ecom_c] != ""].copy() if ecom_c else pd.DataFrame()
    tcs_tech_lower = {t.lower(): t for t in TCS_MAIN_TECHS}
    tcs_nospace = {t.lower().replace(" ",""): t for t in TCS_MAIN_TECHS}

    def map_tech(v):
        import re
        vl = re.sub(r'\s+', ' ', v.lower().strip())
        if not vl: return "Other Technologies"
        if vl in tcs_tech_lower: return tcs_tech_lower[vl]
        vn = vl.replace(" ","")
        if vn in tcs_nospace: return tcs_nospace[vn]
        # Flexible contains match
        if "shopify plus" in vl: return "Shopify Plus"
        if "shopify" in vl and "plus" not in vl: return "Shopify"
        if "bigcommerce" in vl or "big commerce" in vl: return "BigCommerce"
        if "woocommerce" in vl or "woo commerce" in vl: return "WooCommerce"
        if "magento" in vl: return "Magento"
        return "Other Technologies"

    if ecom_c and len(tcs_df)>0:
        tcs_df["_tech"] = tcs_df[ecom_c].apply(map_tech)
    else:
        tcs_df["_tech"] = "Other Technologies"

    tcs_rows = []
    if len(tcs_df) > 0:
        tg = tcs_df.groupby("_tech").agg(leads=(ec,"count"), websites=(wc,"nunique") if wc else (ec,"count")).reset_index()
        tcs_emails = {t: set(g[ec]) for t,g in tcs_df.groupby("_tech")}
        for tech in TCS_MAIN_TECHS + ["Other Technologies"]:
            m = tg[tg["_tech"]==tech]
            if len(m)>0:
                tcs_rows.append({"label":tech,"leads":int(m["leads"].iloc[0]),"websites":int(m["websites"].iloc[0]),"emails":list(tcs_emails.get(tech,set()))})

    # ── Drupal/BinaryWorks ──
    # ── BinaryWorks: rows where Drupal Partners (CMS) has a value ──
    drupal_df = df[df[drupal_c] != ""].copy() if drupal_c else pd.DataFrame()

    def map_cat(v):
        import re
        # Normalize: lowercase, strip, replace all whitespace types with regular space
        vl = re.sub(r'\s+', ' ', v.lower().strip())
        if not vl: return "Other CMS"
        # Drupal versions: match "drupal 7", "drupal7", "drupal 10", etc.
        for ver in ["7","8","9","10","11"]:
            if f"drupal {ver}" in vl or f"drupal{ver}" in vl: return f"Drupal {ver}"
        # Generic Drupal
        if vl == "drupal" or vl == "drupal ": return "Drupal (Generic)"
        # All WordPress variants combined
        if "wordpress" in vl or "word press" in vl: return "WordPress"
        # Everything else
        return "Other CMS"

    DRUPAL_DISPLAY_ORDER = ["Drupal 7","Drupal 8","Drupal 9","Drupal 10","Drupal 11","Drupal (Generic)","WordPress","Other CMS"]

    if drupal_c and len(drupal_df)>0:
        drupal_df["_cat"] = drupal_df[drupal_c].apply(map_cat)
    else:
        drupal_df["_cat"] = "Other CMS"

    drupal_rows = []
    if len(drupal_df) > 0:
        dg = drupal_df.groupby("_cat").agg(leads=(ec,"count"), websites=(wc,"nunique") if wc else (ec,"count")).reset_index()
        drupal_emails = {c: set(g[ec]) for c,g in drupal_df.groupby("_cat")}
        for cat in DRUPAL_DISPLAY_ORDER:
            m = dg[dg["_cat"]==cat]
            if len(m)>0:
                drupal_rows.append({"label":cat,"leads":int(m["leads"].iloc[0]),"websites":int(m["websites"].iloc[0]),"emails":list(drupal_emails.get(cat,set()))})

    # ── ConversionBox ──
    # ── ConversionBox: rows where ConversionBox Competitors has a value ──
    cb_df = df[df[cb_c] != ""].copy() if cb_c else pd.DataFrame()
    cb_rows = []
    if len(cb_df)>0:
        if cb_c:
            cb_df["_comp"] = cb_df[cb_c].replace("","Unspecified")
        else:
            cb_df["_comp"] = "Unspecified"
        cg = cb_df.groupby("_comp").agg(leads=(ec,"count"), websites=(wc,"nunique") if wc else (ec,"count")).reset_index().sort_values("leads",ascending=False)
        cb_emails = {c: set(g[ec]) for c,g in cb_df.groupby("_comp")}
        for _,r in cg.iterrows():
            cb_rows.append({"label":r["_comp"],"leads":int(r["leads"]),"websites":int(r["websites"]),"emails":list(cb_emails.get(r["_comp"],set()))})

    # ── ConversionBox: 200+ Products (banded by Product Count) ──
    # Coerce Product Count to numeric, drop blanks and non-numeric junk, then
    # keep only >= PRODUCT_COUNT_MIN. Each kept contact lands in exactly one band.
    prod_rows = []
    prod_df = pd.DataFrame()
    if prod_c:
        pnum = pd.to_numeric(
            df[prod_c].astype(str).str.replace(",", "", regex=False).str.strip(),
            errors="coerce",
        )
        prod_df = df[pnum >= PRODUCT_COUNT_MIN].copy()
        prod_df["_pc"] = pnum[pnum >= PRODUCT_COUNT_MIN]

        def band_of(v):
            for band in PRODUCT_BANDS:
                if v >= band["lo"] and (band["hi"] is None or v <= band["hi"]):
                    return band["label"]
            return None  # below the floor, already filtered out

        if len(prod_df) > 0:
            prod_df["_band"] = prod_df["_pc"].apply(band_of)
            pg = prod_df.groupby("_band").agg(
                leads=(ec, "count"),
                websites=(wc, "nunique") if wc else (ec, "count"),
            ).reset_index()
            prod_emails = {b: set(g[ec]) for b, g in prod_df.groupby("_band")}
            # Preserve band order rather than group order.
            for band in PRODUCT_BANDS:
                m = pg[pg["_band"] == band["label"]]
                if len(m) > 0:
                    prod_rows.append({
                        "label": band["label"],
                        "leads": int(m["leads"].iloc[0]),
                        "websites": int(m["websites"].iloc[0]),
                        "emails": list(prod_emails.get(band["label"], set())),
                    })

    # ── Others ──
    others = []
    for tag_lower in OTHERS_TAGS:
        odf = df[df["_tag_lower"]==tag_lower]
        if len(odf)>0:
            others.append({"label": OTHERS_DISPLAY.get(tag_lower, tag_lower), "leads":len(odf), "websites":int(odf[wc].nunique()) if wc else 0})

    # ── Overlap ──
    tcs_em = set(tcs_df[ec]) if len(tcs_df)>0 else set()
    drupal_em = set(drupal_df[ec]) if len(drupal_df)>0 else set()
    cb_em = set(cb_df[ec]) if len(cb_df)>0 else set()
    sets = {"tcs":tcs_em,"drupal":drupal_em,"conversionbox":cb_em}
    overlap = {b1:{b2:len(sets[b1]&sets[b2]) for b2 in sets if b2!=b1 and len(sets[b1]&sets[b2])>0} for b1 in sets}

    def totals(rows, brand_df):
        return {"leads": sum(r["leads"] for r in rows), "websites": int(brand_df[wc].nunique()) if wc and len(brand_df)>0 else 0}

    return {
        "tcs": {"rows":tcs_rows, "totals":totals(tcs_rows,tcs_df)},
        "drupal": {"rows":drupal_rows, "totals":totals(drupal_rows,drupal_df)},
        "conversionbox": {"rows":cb_rows, "totals":totals(cb_rows,cb_df)},
        "products": {"rows":prod_rows, "totals":totals(prod_rows,prod_df), "note":""},
        "others": others, "overlap": overlap,
    }


# ════════════════════════════════════════
# LOAD EMAIL MARKETING
# ════════════════════════════════════════
def load_extra_emails():
    """Load the extra email list from Google Drive (emails only, no status)."""
    if not os.path.exists(EXTRA_EMAIL_CSV):
        download_gdrive(EXTRA_EMAIL_GDRIVE_ID, EXTRA_EMAIL_CSV)
    if not os.path.exists(EXTRA_EMAIL_CSV):
        return set()
    try:
        df = pd.read_csv(EXTRA_EMAIL_CSV, low_memory=False, on_bad_lines="skip", encoding="utf-8", encoding_errors="replace")
        email_col = None
        for c in df.columns:
            if c.strip().lower() == "email":
                email_col = c
                break
        if not email_col:
            return set()
        return set(df[email_col].fillna("").astype(str).str.strip()) - {""}
    except:
        return set()


def load_email_mkt(csv_path, brand_rows, extra_emails=None):
    """Load email marketing CSV and combine with extra email list."""
    mkt = {}

    # Load the brand-specific CSV (has opener/non-opener status)
    if os.path.exists(csv_path):
        df = pd.read_csv(csv_path, low_memory=False, on_bad_lines="skip", encoding="utf-8", encoding_errors="replace")
        email_col = status_col = None
        for c in df.columns:
            if c.strip().lower() == "email": email_col = c
            else: status_col = c
        if email_col and status_col:
            df[email_col] = df[email_col].fillna("").astype(str).str.strip()
            df[status_col] = df[status_col].fillna("").astype(str).str.lower()
            for _, row in df.iterrows():
                e = row[email_col]
                if e:
                    is_opener = "opener" in row[status_col] and "non" not in row[status_col]
                    mkt[e] = "opener" if is_opener else "non-opener"

    # Add extra emails (no status info, count as "using" only)
    if extra_emails:
        for e in extra_emails:
            if e and e not in mkt:
                mkt[e] = "using"

    if not mkt:
        return None

    mkt_set = set(mkt.keys())
    result = {"total": 0, "openers": 0, "non_openers": 0, "new_total": 0, "by_tech": {}}
    all_brand_emails = set()
    for tr in brand_rows:
        tech_emails = set(tr.get("emails", []))
        all_brand_emails.update(tech_emails)
        matched = tech_emails & mkt_set
        op = sum(1 for e in matched if mkt.get(e) == "opener")
        non_op = sum(1 for e in matched if mkt.get(e) == "non-opener")
        result["by_tech"][tr["label"]] = {
            "total": len(matched), "openers": op,
            "non_openers": non_op,
            "new_emails": len(tech_emails) - len(matched)
        }
        result["total"] += len(matched)
        result["openers"] += op
        result["non_openers"] += non_op
    result["new_total"] = len(all_brand_emails) - len(all_brand_emails & mkt_set)
    return result


# ════════════════════════════════════════
# PARSE CONTRIBUTION SHEETS
# ════════════════════════════════════════
def read_gsheet_tab(gid):
    """Read a Google Sheet tab as a pandas DataFrame using published URL."""
    url = f"https://docs.google.com/spreadsheets/d/e/{GSHEET_PUB_KEY}/pub?gid={gid}&single=true&output=csv"
    try:
        df = pd.read_csv(url, on_bad_lines="skip", encoding="utf-8", encoding_errors="replace")
        if len(df) > 0:
            return df
    except:
        pass
    return None


def parse_contribution():
    """Read individual contribution data from Google Sheets."""
    persons = {}
    sales_data = {}

    PERSON_DISPLAY = {"kishore": "Kishore", "ilakkiya": "Ilakkiya", "dharanshri": "Dharanshri"}

    for key, tab_info in GSHEET_TABS.items():
        if key == "sales":
            continue  # Handle sales separately
        df = read_gsheet_tab(tab_info["gid"])
        if df is None or len(df) == 0:
            continue

        # Normalize column names
        df.columns = [str(c).strip().lower() for c in df.columns]

        # Find columns
        month_col = count_col = cat_col = date_col = None
        for c in df.columns:
            if c == "month": month_col = c
            elif c in ("count", "lead count"): count_col = c
            elif c in ("technology", "domain name", "domain", "category"): cat_col = c
            elif c in ("date", "date "): date_col = c

        # Fallback: find count column
        if count_col is None:
            for c in df.columns:
                if df[c].dtype in ("int64", "float64"):
                    count_col = c
                    break

        if count_col is None:
            continue

        months = OrderedDict()

        for _, row in df.iterrows():
            # Get month
            month_key = None
            date_str = ""
            if month_col and pd.notna(row.get(month_col)):
                mv = str(row[month_col]).strip()
                for mn in MONTH_ORDER:
                    if mv.lower().startswith(mn):
                        month_key = mn.capitalize()
                        break
                date_str = month_key or mv
            elif date_col and pd.notna(row.get(date_col)):
                dv_str = str(row[date_col]).strip()
                # Try multiple date formats
                month_key = None
                date_str = ""
                # First: check if it starts with a month name (e.g. "Jan 2026", "Jan")
                for mn in MONTH_ORDER:
                    if dv_str.lower().startswith(mn):
                        month_key = mn.capitalize()
                        date_str = month_key
                        break
                # Second: try parsing as full date (e.g. "23 Jan 2026", "2026-01-23")
                if not month_key:
                    try:
                        dv = pd.to_datetime(dv_str, dayfirst=True)
                        month_key = dv.strftime("%b")
                        date_str = dv.strftime("%b %d")
                    except:
                        pass
                # Third: check if day comes first (e.g. "23 Jan 2026")
                if not month_key:
                    parts = dv_str.split()
                    for p in parts:
                        for mn in MONTH_ORDER:
                            if p.lower().startswith(mn):
                                month_key = mn.capitalize()
                                date_str = dv_str
                                break
                        if month_key: break

            if not month_key:
                continue

            # Get count (handle commas like "4,335")
            try:
                count = int(float(str(row[count_col]).replace(",", "")))
            except:
                continue
            if count <= 0:
                continue

            # Get category
            category = "General"
            if cat_col and pd.notna(row.get(cat_col)):
                category = str(row[cat_col]).strip()

            if month_key not in months:
                months[month_key] = []
            months[month_key].append({"date": date_str, "category": category, "count": count})

        # Build output sorted newest first
        person_months = []
        for mk in sorted(months.keys(), key=lambda m: MONTH_ORDER.get(m.lower()[:3], 0), reverse=True):
            entries = months[mk]
            if entries:
                person_months.append({"month": f"{mk} 2026", "total": sum(e["count"] for e in entries), "entries": entries})

        if person_months:
            display_name = tab_info["name"]
            persons[key] = {"name": display_name, "months": person_months}

    # Parse sales tab
    sales_gid = GSHEET_TABS.get("sales", {}).get("gid")
    sales_df = read_gsheet_tab(sales_gid) if sales_gid else None
    if sales_df is not None and len(sales_df) > 0:
        sales_df.columns = [str(c).strip().lower() for c in sales_df.columns]
        name_col = count_col = month_col = tech_col = None
        for c in sales_df.columns:
            if c == "name": name_col = c
            elif c == "month": month_col = c
            elif c in ("count",): count_col = c
            elif c in ("technology",): tech_col = c

        if name_col and count_col and month_col:
            for _, row in sales_df.iterrows():
                name = str(row[name_col]).strip().lower() if pd.notna(row.get(name_col)) else ""
                month_raw = str(row[month_col]).strip() if pd.notna(row.get(month_col)) else ""
                try:
                    count = int(float(row[count_col]))
                except:
                    continue

                month_key = None
                for mn in MONTH_ORDER:
                    if month_raw.lower().startswith(mn):
                        month_key = mn.capitalize()
                        break
                if not month_key:
                    continue

                # Normalize name
                name_map = {"dharanishri": "dharanshri", "dharanshri": "dharanshri",
                           "kishore": "kishore", "ilakkiya": "ilakkiya", "illakkia": "ilakkiya"}
                for variant, normalized in name_map.items():
                    if variant in name:
                        name = normalized
                        break

                if name not in sales_data:
                    sales_data[name] = {}
                if month_key not in sales_data[name]:
                    sales_data[name][month_key] = 0
                sales_data[name][month_key] += count

    return {"persons": persons, "sales": sales_data}


# ════════════════════════════════════════
# LOAD ALL DATA
# ════════════════════════════════════════
def load_all():
    data = {
        "brands": {
            "tcs": {"name":"CommerceShop","colLabel":"Technology","rows":[],"totals":{"leads":0,"websites":0}},
            "drupal": {"name":"BinaryWorks","colLabel":"Category","rows":[],"totals":{"leads":0,"websites":0}},
            "conversionbox": {"name":"ConversionBox","colLabel":"Technology","rows":[],"totals":{"leads":0,"websites":0}},
        },
        # products: the ConversionBox "200+ Products" section, rendered under the
        # competitors breakdown on the ConversionBox tab.
        "products": {"rows": [], "totals": {"leads": 0, "websites": 0}, "note": ""},
        "others": [], "overlap": {"tcs":{},"drupal":{},"conversionbox":{}},
        "email_mkt": {"tcs":None,"drupal":None,"conversionbox":None}, "persons": {}, "sales": {},
    }

    # Download HubSpot from Google Drive
    if not os.path.exists(HUBSPOT_CSV):
        with st.spinner("Downloading data from Google Drive..."):
            download_gdrive(HUBSPOT_GDRIVE_ID, HUBSPOT_CSV)

    hub = load_hubspot()
    if hub:
        for b in ["tcs","drupal","conversionbox"]:
            data["brands"][b]["rows"] = hub[b]["rows"]
            data["brands"][b]["totals"] = hub[b]["totals"]
        data["products"] = hub.get("products", data["products"])
        data["others"] = hub["others"]
        data["overlap"] = hub["overlap"]
        # Load extra email list (covers all brands)
        extra_emails = load_extra_emails()

        data["email_mkt"]["tcs"] = load_email_mkt(TCS_EMAIL_MKT_CSV, hub["tcs"]["rows"], extra_emails)
        data["email_mkt"]["drupal"] = load_email_mkt(BW_EMAIL_MKT_CSV, hub["drupal"]["rows"], extra_emails)
        # ConversionBox: no brand-specific CSV, but use extra emails
        cb_mkt = load_email_mkt("__nonexistent__", hub["conversionbox"]["rows"], extra_emails)
        if cb_mkt:
            data["email_mkt"]["conversionbox"] = cb_mkt

    # Option A: the Product Count column is not in the Drive CSV, so the 200+
    # Products section is pulled LIVE from HubSpot on every load. Cached so a
    # normal Streamlit rerun (any button click) does not refire the 4 API calls.
    live_products = cached_product_bands()
    if live_products and live_products["rows"]:
        data["products"] = live_products

    contrib = parse_contribution()
    data["persons"] = contrib.get("persons", {})
    data["sales"] = contrib.get("sales", {})

    # Fallback: if no persons from Google Sheets, try xlsx file
    if not data["persons"]:
        for fname in ["Copy of Over all DB .xlsx", "Copy of Over all DB.xlsx", "Copy_of_Over_all_DB.xlsx"]:
            p = os.path.join(APP_DIR, fname)
            if os.path.exists(p):
                try:
                    wb = openpyxl.load_workbook(p, data_only=True)
                    # Simple parse from xlsx as fallback
                    for sn in wb.sheetnames:
                        pname = None
                        for check in ["kishore","ilakkiya","illakkia","dharanshri","dharanishri"]:
                            if check in sn.lower():
                                pname = check.capitalize()
                                if pname == "Illakkia": pname = "Ilakkiya"
                                if pname == "Dharanishri": pname = "Dharanshri"
                                break
                        if not pname: continue
                        ws = wb[sn]
                        rows_data = [list(r) for r in ws.iter_rows(min_row=1, values_only=True)]
                        if len(rows_data) < 2: continue
                        # Simple: assume Month col 0, Tech col 1 or 2, Count col 2 or 3
                        months_d = OrderedDict()
                        for row in rows_data[1:]:
                            try:
                                mv = str(row[0]).strip() if row[0] else ""
                                mk = None
                                for mn in MONTH_ORDER:
                                    if mv.lower().startswith(mn): mk = mn.capitalize(); break
                                if not mk and isinstance(row[0], datetime): mk = row[0].strftime("%b")
                                if not mk: continue
                                # Find count (first numeric value)
                                count = None
                                cat = "General"
                                for i, v in enumerate(row[1:], 1):
                                    if v is None: continue
                                    try:
                                        count = int(float(v)); break
                                    except:
                                        if not cat or cat == "General": cat = str(v).strip()
                                if not count or count <= 0: continue
                                if mk not in months_d: months_d[mk] = []
                                months_d[mk].append({"date": mk, "category": cat, "count": count})
                            except: continue
                        pm = [{"month": f"{m} 2026", "total": sum(e["count"] for e in months_d[m]), "entries": months_d[m]}
                              for m in sorted(months_d.keys(), key=lambda x: MONTH_ORDER.get(x.lower()[:3],0), reverse=True)]
                        if pm: data["persons"][pname.lower()] = {"name": pname, "months": pm}
                    wb.close()
                except: pass
                break

    return data


# ════════════════════════════════════════
# HTML DASHBOARD
# ════════════════════════════════════════
def build_html(data):
    bc = {}
    for bk,bv in data["brands"].items():
        bc[bk] = {"name":bv["name"],"colLabel":bv["colLabel"],
                  "rows":[{"label":r["label"],"leads":r["leads"],"websites":r["websites"]} for r in bv["rows"]],
                  "totals":bv["totals"]}
    bj = json.dumps(bc)
    oj = json.dumps(data["others"])
    prod = data.get("products", {"rows": [], "totals": {"leads": 0, "websites": 0}, "note": ""})
    prodj = json.dumps({
        "rows": [{"label": r["label"], "leads": r["leads"], "websites": r["websites"]} for r in prod["rows"]],
        "totals": prod["totals"],
        "note": prod.get("note", ""),
    })
    ovj = json.dumps(data["overlap"])
    emj = json.dumps({k: {"total":v["total"],"openers":v["openers"],"non_openers":v["non_openers"],
          "new_total":v["new_total"],"by_tech":v["by_tech"]} if v else None for k,v in data["email_mkt"].items()})
    pj = json.dumps(data["persons"])
    sj = json.dumps(data["sales"])

    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
*{{margin:0;padding:0;box-sizing:border-box}}
:root{{--bg:#f8f9fa;--bg-card:#fff;--bg-el:#f1f3f5;--bdr:#e0e0e0;--bdr-l:#eee;
--txt:#1a1a1a;--txt-s:#666;--txt-m:#999;--accent:#4338ED;--accent2:#F97316;
--glow:rgba(67,56,237,0.15);--grad:linear-gradient(135deg,#4338ED,#6366F1);
--grad-bar:linear-gradient(90deg,#4338ED,#818CF8);--stat-bg:#fff;--stat-bdr:#e0e0e0}}
body{{font-family:'Inter',sans-serif;background:var(--bg);color:var(--txt);line-height:1.5}}
.db{{max-width:1180px;margin:0 auto;padding:36px 24px 56px}}
.hdr{{display:flex;align-items:center;justify-content:space-between;margin-bottom:32px;flex-wrap:wrap;gap:16px}}
.hdr h1{{font-size:26px;font-weight:900}}.hdr h1 .hl{{color:var(--accent)}}
.hdr .sub{{font-size:13px;color:var(--txt-s);margin-top:3px}}
.tg{{display:inline-flex;background:var(--bg-card);border-radius:12px;padding:4px;border:1px solid var(--bdr)}}
.tb{{padding:9px 22px;border:none;background:transparent;font-family:inherit;font-size:13px;font-weight:600;color:var(--txt-s);border-radius:9px;cursor:pointer;transition:all .25s;white-space:nowrap}}
.tb:hover{{color:var(--txt);background:var(--bg-el)}}
.tb.on{{background:var(--accent);color:#fff;box-shadow:0 4px 16px var(--glow)}}
.sr{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;margin-bottom:28px}}
.sc{{background:var(--stat-bg);border-radius:16px;padding:22px 24px;border:1px solid var(--stat-bdr);position:relative;overflow:hidden}}
.sc::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:var(--grad)}}
.sc .lb{{font-size:10px;font-weight:700;color:var(--txt-s);text-transform:uppercase;letter-spacing:.7px;margin-bottom:8px}}
.sc .vl{{font-size:30px;font-weight:900;color:var(--txt)}}
.sc .ic{{position:absolute;top:18px;right:20px;width:36px;height:36px;border-radius:10px;background:rgba(99,102,241,.1);display:flex;align-items:center;justify-content:center;font-size:15px}}
.pn{{background:var(--bg-card);border-radius:18px;padding:28px;border:1px solid var(--bdr);margin-bottom:28px}}
.pt{{font-size:17px;font-weight:800;color:var(--txt);margin-bottom:22px}}
.bc{{display:flex;flex-direction:column;gap:10px}}
.br{{display:grid;grid-template-columns:150px 1fr 80px 80px;align-items:center;gap:14px;padding:5px 0;border-radius:8px;transition:background .15s}}
.br.hm{{grid-template-columns:150px 1fr 80px 80px 90px 90px}}
.br.pcol{{grid-template-columns:150px 1fr 90px}}
.br:hover{{background:rgba(0,0,0,.02)}}
.bl{{font-size:13px;font-weight:600;color:var(--txt);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.bt{{height:30px;background:var(--bg-el);border-radius:8px;overflow:hidden}}
.bf{{height:100%;border-radius:8px;transition:width .7s cubic-bezier(.22,1,.36,1);min-width:6px;position:relative}}
.bf::after{{content:'';position:absolute;top:0;right:0;width:40px;height:100%;background:linear-gradient(90deg,transparent,rgba(255,255,255,.18));border-radius:0 8px 8px 0}}
.bv{{font-size:13px;font-weight:800;color:var(--txt);text-align:right;font-variant-numeric:tabular-nums}}
.bv.sec{{color:var(--txt-s);font-weight:600}}.bv.mkt{{color:var(--accent2);font-weight:700}}.bv.new{{color:#10B981;font-weight:700}}
.bch{{font-size:10px;font-weight:700;color:var(--txt-m);text-transform:uppercase;letter-spacing:.5px;text-align:right}}.bch.l{{text-align:left}}
.tr{{display:grid;grid-template-columns:150px 1fr 80px 80px;align-items:center;gap:14px;margin-top:16px;padding-top:16px;border-top:2px solid var(--accent)}}
.tr.hm{{grid-template-columns:150px 1fr 80px 80px 90px 90px}}
.tr.pcol{{grid-template-columns:150px 1fr 90px}}
.tr .bl{{font-weight:900;color:var(--accent);font-size:14px}}.tr .bv{{font-weight:900;color:var(--accent)}}.tr .bv.sec{{color:var(--accent2)}}
.olt{{display:inline-block;padding:2px 8px;border-radius:5px;font-size:9px;font-weight:700;background:rgba(249,115,22,.12);color:#F97316;border:1px solid rgba(249,115,22,.3);margin-left:6px}}
.mt{{display:inline-flex;background:var(--bg-el);border-radius:8px;padding:2px;border:1px solid var(--bdr);margin-left:12px}}
.mt button{{padding:4px 10px;border:none;background:transparent;font-family:inherit;font-size:10px;font-weight:600;color:var(--txt-m);border-radius:6px;cursor:pointer;transition:all .2s}}.mt button.on{{background:var(--accent2);color:#fff}}
.oc-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px}}
.oc{{background:var(--stat-bg);border-radius:16px;padding:24px;border:1px solid var(--stat-bdr);position:relative;overflow:hidden}}
.oc::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:var(--grad)}}
.oc .on{{font-size:14px;font-weight:800;color:var(--txt);margin-bottom:12px}}
.oc .or{{display:flex;justify-content:space-between;margin-bottom:6px}}.oc .ol{{font-size:11px;color:var(--txt-s)}}.oc .ov{{font-size:14px;font-weight:800}}
.cw{{display:flex;justify-content:center;margin:44px 0 28px}}
.cb{{display:inline-flex;align-items:center;gap:10px;padding:14px 40px;background:var(--bg-card);border:1px solid var(--bdr);border-radius:14px;font-family:inherit;font-size:15px;font-weight:700;color:var(--txt);cursor:pointer;transition:all .25s}}
.cb:hover{{border-color:var(--accent);box-shadow:0 0 24px var(--glow)}}.cb.on{{background:var(--accent);color:#fff;border-color:transparent}}
.cb .ar{{display:inline-block;transition:transform .3s;font-size:11px}}.cb.on .ar{{transform:rotate(180deg)}}
.cs{{display:none}}.cs.op{{display:block;animation:fi .35s ease}}
@keyframes fi{{from{{opacity:0;transform:translateY(-12px)}}to{{opacity:1;transform:translateY(0)}}}}
.pw{{display:flex;justify-content:center;margin-bottom:28px}}
.gp{{background:var(--bg-card);border-radius:18px;padding:28px;border:1px solid var(--bdr);margin-bottom:24px}}
.gt{{font-size:15px;font-weight:800;margin-bottom:24px}}.ga{{position:relative;height:240px;display:flex;align-items:flex-end;padding-bottom:36px;padding-left:60px}}
.gy{{position:absolute;left:0;top:0;bottom:36px;width:55px;display:flex;flex-direction:column;justify-content:space-between;align-items:flex-end;padding-right:8px}}
.yl{{font-size:10px;color:var(--txt-m);font-variant-numeric:tabular-nums;font-weight:600}}
.gg{{position:absolute;left:60px;right:0;top:0;bottom:36px}}.gl{{position:absolute;left:0;right:0;height:1px;background:var(--bdr-l);opacity:.5}}
.gbs{{display:flex;align-items:flex-end;gap:16px;flex:1;height:calc(100% - 36px);position:relative;z-index:1}}
.gbg{{flex:1;display:flex;flex-direction:column;align-items:center;height:100%;justify-content:flex-end;position:relative}}
.gb{{width:100%;max-width:64px;border-radius:10px 10px 4px 4px;position:relative;transition:all .5s cubic-bezier(.22,1,.36,1);cursor:pointer}}
.gb:hover{{filter:brightness(1.1)}}.gb .gv{{position:absolute;top:-24px;left:50%;transform:translateX(-50%);font-size:12px;font-weight:800;white-space:nowrap}}
.gbl{{position:absolute;bottom:-30px;left:50%;transform:translateX(-50%);font-size:11px;font-weight:700;color:var(--txt-s);white-space:nowrap}}
.pp{{background:var(--bg-card);border-radius:18px;border:1px solid var(--bdr);margin-bottom:20px;overflow:hidden;display:none}}.pp.vis{{display:block}}
.pph{{padding:18px 24px;border-bottom:1px solid var(--bdr);background:var(--bg-el)}}.ppn{{font-size:17px;font-weight:800}}
.mr{{border-bottom:1px solid var(--bdr-l)}}.mr:last-child{{border-bottom:none}}
.mh{{display:flex;align-items:center;justify-content:space-between;padding:14px 24px;cursor:pointer;transition:background .15s;user-select:none}}
.mh:hover{{background:var(--bg-el)}}.mh .ml{{display:flex;align-items:center;gap:10px}}
.mh .ma{{font-size:11px;color:var(--txt-m);transition:transform .2s;display:inline-block;width:16px;text-align:center}}.mh .ma.op{{transform:rotate(90deg);color:var(--accent)}}
.mh .mm{{font-size:14px;font-weight:700}}.mh .mc{{font-size:15px;font-weight:800;font-variant-numeric:tabular-nums}}
.mh .me{{font-size:11px;color:var(--txt-m);margin-left:8px}}
.md{{display:none;background:#f8f9fa;border-top:1px solid var(--bdr-l)}}.md.op{{display:block}}
.dr{{display:grid;grid-template-columns:100px 1fr 100px;padding:11px 24px 11px 50px;gap:12px;border-bottom:1px solid #eee;align-items:center}}.dr:last-child{{border-bottom:none}}
.dd{{font-size:12px;color:var(--txt-s);font-weight:600}}.dt{{display:inline-block;padding:3px 10px;border-radius:6px;font-size:10px;font-weight:700}}
.dc{{font-size:14px;font-weight:800;text-align:right;font-variant-numeric:tabular-nums}}
.msw{{display:flex;justify-content:space-between;padding:11px 24px 11px 50px;background:var(--bg-el);border-top:1px solid var(--bdr)}}.msw span{{font-size:13px;font-weight:800;color:var(--accent)}}
.nd{{text-align:center;padding:48px;color:var(--txt-m);font-size:14px}}.ft{{text-align:center;padding:28px 0;font-size:11px;color:var(--txt-m)}}
@media(max-width:768px){{.db{{padding:16px 12px}}.hdr{{flex-direction:column;align-items:flex-start}}.sr{{grid-template-columns:1fr 1fr}}
.br,.tr{{grid-template-columns:100px 1fr 60px 60px;gap:8px}}.br.hm,.tr.hm{{grid-template-columns:100px 1fr 60px 60px 60px 60px}}.pn,.gp{{padding:20px 16px}}}}
</style></head><body>
<div class="db" id="app"></div>
<script>
const TH={{tcs:{{a:'#4338ED',a2:'#F97316',g:'rgba(67,56,237,.12)',gb:'linear-gradient(90deg,#4338ED,#818CF8)',c:['#4338ED','#818CF8']}},
drupal:{{a:'#8B5CF6',a2:'#F59E0B',g:'rgba(139,92,246,.12)',gb:'linear-gradient(90deg,#8B5CF6,#C4B5FD)',c:['#8B5CF6','#C4B5FD']}},
conversionbox:{{a:'#2563EB',a2:'#10B981',g:'rgba(37,99,235,.12)',gb:'linear-gradient(90deg,#2563EB,#60A5FA)',c:['#2563EB','#60A5FA']}},
others:{{a:'#6B7280',a2:'#9CA3AF',g:'rgba(107,114,128,.12)',gb:'linear-gradient(90deg,#6B7280,#9CA3AF)',c:['#6B7280','#9CA3AF']}}}};
const B={bj};const OT={oj};const OV={ovj};const EM={emj};const P={pj};const SL={sj};const PRODUCTS={prodj};
let aB='tcs',cO=false,aP=Object.keys(P)[0]||'',oM={{}},mF='all';
function fmt(n){{return n.toLocaleString('en-IN')}}
function th(t){{const s=document.documentElement.style,h=TH[t]||TH.others;s.setProperty('--accent',h.a);s.setProperty('--accent2',h.a2);s.setProperty('--glow',h.g);s.setProperty('--grad-bar',h.gb);s.setProperty('--grad','linear-gradient(135deg,'+h.a+','+h.a+'aa)')}}
function cc(c){{return{{'ConversionBox':'#6366F1','Drupal':'#3B82F6','TCS':'#4338ED','Consultant':'#8B5CF6','Freelancer':'#A855F7','ABM-Fintech':'#06B6D4','ABM- Higher edu':'#0891B2','Tidio':'#10B981','Manufacturing List':'#F59E0B','TCS New Lead List':'#3B82F6','TCS New lead List':'#3B82F6','WordPress':'#0EA5E9','Other\\'s CMS':'#F97316','NGO':'#EC4899','General':'#6B7280'}}[c]||'#6B7280'}}
function sB(k){{aB=k;mF='all';render()}}function tC(){{cO=!cO;render()}}function sP(k){{aP=k;render()}}function tM(p,m){{const k=p+'-'+m;oM[k]=!oM[k];render()}}function sM(f){{mF=f;render()}}
function chart(p,t){{const ms=[...p.months].reverse(),mx=Math.max(...ms.map(m=>m.total)),st=4,sv=Math.ceil(mx/st/1000)*1000,cl=sv*st;
let g='',y='';for(let i=0;i<=st;i++){{const pc=(i/st)*100,v=sv*i;g+=`<div class="gl" style="bottom:${{pc}}%"></div>`;y+=`<div class="yl">${{v>=1000?(v/1000).toFixed(0)+'K':v}}</div>`}}
// Target lines: 3K red, 5K yellow, 10K green (behind bars)
const targets=[{{v:3000,c:'#EF4444',l:'3K'}},{{v:5000,c:'#F59E0B',l:'5K'}},{{v:10000,c:'#10B981',l:'10K'}}];
let tl='';targets.forEach(t=>{{if(t.v<=cl){{const tp=(t.v/cl)*100;tl+=`<div style="position:absolute;left:0;right:0;bottom:${{tp}}%;height:1px;border-top:2px dashed ${{t.c}};opacity:0.4;z-index:0"></div><div style="position:absolute;left:-35px;bottom:${{tp-1}}%;font-size:8px;font-weight:700;color:${{t.c}};z-index:0">${{t.l}}</div>`}}}});
const[c1,c2]=t.c;const bs=ms.map(m=>{{const h=cl>0?(m.total/cl)*100:0;return`<div class="gbg"><div class="gb" style="height:${{Math.max(h,3)}}%;background:linear-gradient(180deg,${{c2}},${{c1}})"><div class="gv">${{fmt(m.total)}}</div><div class="gbl">${{m.month.split(' ')[0]}}</div></div></div>`}}).join('');
return`<div class="gp"><div class="gt">${{p.name}}'s Monthly Contribution</div><div class="ga"><div class="gy">${{y}}</div><div class="gg">${{g}}${{tl}}</div><div class="gbs">${{bs}}</div></div></div>`}}
function render(){{const b=B[aB],t=TH[aB]||TH.others;th(aB);const isO=aB==='others';
const ov=OV[aB]||{{}};let ovH='';for(const[ob,cnt] of Object.entries(ov)){{const nm={{tcs:'TCS',drupal:'BinaryWorks',conversionbox:'ConversionBox'}};ovH+=`<span class="olt">${{fmt(cnt)}} also in ${{nm[ob]||ob}}</span>`}}
const mkt=EM[aB]||null;const hM=!!mkt;const mc=hM?'hm':'';
let main='';
if(isO){{const st=`<div class="sr"><div class="sc"><div class="lb">Total tags</div><div class="vl">${{OT.length}}</div><div class="ic">🏷️</div></div><div class="sc"><div class="lb">Total leads</div><div class="vl">${{fmt(OT.reduce((s,o)=>s+o.leads,0))}}</div><div class="ic">📊</div></div><div class="sc"><div class="lb">Total websites</div><div class="vl">${{fmt(OT.reduce((s,o)=>s+o.websites,0))}}</div><div class="ic">🌐</div></div></div>`;
const cd=OT.length?`<div class="oc-grid">${{OT.map(o=>`<div class="oc"><div class="on">${{o.label}}</div><div class="or"><span class="ol">Leads</span><span class="ov">${{fmt(o.leads)}}</span></div><div class="or"><span class="ol">Websites</span><span class="ov">${{fmt(o.websites)}}</span></div></div>`).join('')}}</div>`:'<div class="nd">No data</div>';
main=st+cd}}else{{if(!b||!b.rows.length){{main=`<div class="sr"><div class="sc"><div class="lb">Total leads</div><div class="vl">0</div></div></div><div class="nd">No data found for this brand.</div>`}}else{{
const mx=Math.max(...b.rows.map(r=>r.leads));
let st=`<div class="sr"><div class="sc"><div class="lb">Total leads</div><div class="vl">${{fmt(b.totals.leads)}}</div><div class="ic">📊</div></div><div class="sc"><div class="lb">Unique websites</div><div class="vl">${{fmt(b.totals.websites)}}</div><div class="ic">🌐</div></div><div class="sc"><div class="lb">Avg leads / website</div><div class="vl">${{(b.totals.leads/(b.totals.websites||1)).toFixed(1)}}</div><div class="ic">⚡</div></div>`;
if(hM)st+=`<div class="sc"><div class="lb">Email mkt using</div><div class="vl">${{fmt(mkt.total)}}</div><div class="ic">📧</div></div><div class="sc"><div class="lb">New emails</div><div class="vl">${{fmt(mkt.new_total)}}</div><div class="ic">✨</div></div>`;
st+=`</div>`;if(ovH)st+=`<div style="margin:-16px 0 20px">${{ovH}}</div>`;
let mtg='';if(hM)mtg=`<div class="mt"><button class="${{mF==='all'?'on':''}}" onclick="sM('all')">All</button><button class="${{mF==='opener'?'on':''}}" onclick="sM('opener')">Openers</button><button class="${{mF==='non-opener'?'on':''}}" onclick="sM('non-opener')">Non-Openers</button></div>`;
let hd=`<div class="br ${{mc}}" style="margin-bottom:4px"><div class="bch l">${{b.colLabel}}</div><div></div><div class="bch">Leads</div><div class="bch">Sites</div>`;if(hM)hd+=`<div class="bch">Mkt</div><div class="bch">New</div>`;hd+=`</div>`;
let rw=b.rows.map(r=>{{let mv='',nv='';if(hM&&mkt.by_tech[r.label]){{const mt=mkt.by_tech[r.label];mv=fmt(mF==='all'?mt.total:mF==='opener'?mt.openers:mt.non_openers);nv=fmt(mt.new_emails)}}return`<div class="br ${{mc}}"><div class="bl">${{r.label}}</div><div class="bt"><div class="bf" style="width:${{(r.leads/mx*100).toFixed(1)}}%;background:var(--grad-bar)"></div></div><div class="bv">${{fmt(r.leads)}}</div><div class="bv sec">${{fmt(r.websites)}}</div>${{hM?`<div class="bv mkt">${{mv}}</div><div class="bv new">${{nv}}</div>`:''}} </div>`}}).join('');
let tm='',tn='';if(hM){{tm=fmt(mF==='all'?mkt.total:mF==='opener'?mkt.openers:mkt.non_openers);tn=fmt(mkt.new_total)}}
let tot=`<div class="tr ${{mc}}"><div class="bl">TOTAL</div><div></div><div class="bv">${{fmt(b.totals.leads)}}</div><div class="bv sec">${{fmt(b.totals.websites)}}</div>${{hM?`<div class="bv mkt">${{tm}}</div><div class="bv new">${{tn}}</div>`:''}}</div>`;
main=st+`<div class="pn"><div style="display:flex;align-items:center;flex-wrap:wrap;margin-bottom:22px"><div class="pt" style="margin:0">${{b.colLabel}} Breakdown</div>${{mtg}}</div><div class="bc">${{hd}}${{rw}}${{tot}}</div></div>`;
// ConversionBox: competitors on the LEFT, "200+ Products" on the RIGHT, 50-50.
// Products are pulled live from HubSpot (Option A); the CSV has no Product
// Count column. Sub-200 and blanks are excluded upstream. The products panel
// has NO Sites column: the HubSpot count API returns totals only, so a Sites
// number would always be a meaningless 0.
if(aB==='conversionbox'){{
const hasProd = PRODUCTS && PRODUCTS.rows && PRODUCTS.rows.length && PRODUCTS.totals.leads>0;

// Left panel: the competitors breakdown already built above as `main`, minus
// the leading stat cards (st) which stay at the top spanning both columns.
const leftPanel = `<div class="pn" style="margin:0"><div style="display:flex;align-items:center;flex-wrap:wrap;margin-bottom:22px"><div class="pt" style="margin:0">${{b.colLabel}} Breakdown</div>${{mtg}}</div><div class="bc">${{hd}}${{rw}}${{tot}}</div></div>`;

let rightPanel;
if(hasProd){{
// 3-column grid (label, bar, leads): no Sites column.
const pr=PRODUCTS.rows;const pMax=Math.max(...pr.map(r=>r.leads),1);
const pHead=`<div class="br pcol" style="margin-bottom:4px"><div class="bch l">Product Count</div><div></div><div class="bch">Leads</div></div>`;
const pRows=pr.map(r=>`<div class="br pcol"><div class="bl">${{r.label}}</div><div class="bt"><div class="bf" style="width:${{(r.leads/pMax*100).toFixed(1)}}%;background:var(--grad-bar)"></div></div><div class="bv">${{fmt(r.leads)}}</div></div>`).join('');
const pTot=`<div class="tr pcol"><div class="bl">TOTAL</div><div></div><div class="bv">${{fmt(PRODUCTS.totals.leads)}}</div></div>`;
rightPanel=`<div class="pn" style="margin:0"><div class="pt">&#128230; 200+ Products <span style="font-size:12px;font-weight:600;color:var(--txt-m)">(live, under 200 excluded)</span></div><div class="bc">${{pHead}}${{pRows}}${{pTot}}</div></div>`;
}}else{{
const why = (PRODUCTS && PRODUCTS.note) ? PRODUCTS.note : 'no data returned';
rightPanel=`<div class="pn" style="margin:0"><div class="pt">&#128230; 200+ Products <span style="font-size:12px;font-weight:600;color:var(--txt-m)">(live from HubSpot)</span></div><div class="nd" style="text-align:left;padding:20px 4px">No Product Count data loaded.<br><span style="font-size:12px;color:var(--txt-m)">Resolver: ${{why}}</span><br><span style="font-size:12px;color:var(--txt-m)">If this says NOT FOUND, the Product Count property name is different, or the HubSpot key is missing in Streamlit Secrets.</span></div></div>`;
}}

// Replace the single stacked panel with the two side-by-side panels.
main=st+`<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:28px">${{leftPanel}}${{rightPanel}}</div>`;
}}
// BinaryWorks split: show Drupal and WordPress side by side
if(aB==='drupal'){{
const drupalRows=b.rows.filter(r=>r.label.toLowerCase().startsWith('drupal'));
const wpRows=b.rows.filter(r=>r.label.toLowerCase().startsWith('wordpress'));
const otherRows=b.rows.filter(r=>!r.label.toLowerCase().startsWith('drupal')&&!r.label.toLowerCase().startsWith('wordpress'));
const dTotal=drupalRows.reduce((s,r)=>s+r.leads,0);const wTotal=wpRows.reduce((s,r)=>s+r.leads,0);const oTotal=otherRows.reduce((s,r)=>s+r.leads,0);
const dMax=Math.max(...drupalRows.map(r=>r.leads),1);const wMax=Math.max(...wpRows.map(r=>r.leads),1);
function mkRows(rows,mxL){{return rows.map(r=>{{let mv='',nv='';if(hM&&mkt.by_tech[r.label]){{const mt=mkt.by_tech[r.label];mv=fmt(mF==='all'?mt.total:mF==='opener'?mt.openers:mt.non_openers);nv=fmt(mt.new_emails)}}return`<div class="br ${{mc}}"><div class="bl">${{r.label}}</div><div class="bt"><div class="bf" style="width:${{(r.leads/mxL*100).toFixed(1)}}%;background:var(--grad-bar)"></div></div><div class="bv">${{fmt(r.leads)}}</div><div class="bv sec">${{fmt(r.websites)}}</div>${{hM?`<div class="bv mkt">${{mv}}</div><div class="bv new">${{nv}}</div>`:''}} </div>`}}).join('')}}
const dTotMkt=hM?drupalRows.reduce((s,r)=>{{const mt=mkt.by_tech[r.label];return s+(mt?mt.total:0)}},0):0;
const wTotMkt=hM?wpRows.reduce((s,r)=>{{const mt=mkt.by_tech[r.label];return s+(mt?mt.total:0)}},0):0;
const dTotNew=hM?drupalRows.reduce((s,r)=>{{const mt=mkt.by_tech[r.label];return s+(mt?mt.new_emails:0)}},0):0;
const wTotNew=hM?wpRows.reduce((s,r)=>{{const mt=mkt.by_tech[r.label];return s+(mt?mt.new_emails:0)}},0):0;
main=st+`<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:28px">
<div class="pn" style="margin:0"><div style="display:flex;align-items:center;flex-wrap:wrap;margin-bottom:18px"><div class="pt" style="margin:0;color:#8B5CF6">🔮 Drupal</div>${{mtg}}</div><div class="bc">${{hd}}${{mkRows(drupalRows,dMax)}}<div class="tr ${{mc}}"><div class="bl">SUBTOTAL</div><div></div><div class="bv">${{fmt(dTotal)}}</div><div class="bv sec">${{fmt(drupalRows.reduce((s,r)=>s+r.websites,0))}}</div>${{hM?`<div class="bv mkt">${{fmt(dTotMkt)}}</div><div class="bv new">${{fmt(dTotNew)}}</div>`:''}}</div></div></div>
<div class="pn" style="margin:0"><div style="display:flex;align-items:center;flex-wrap:wrap;margin-bottom:18px"><div class="pt" style="margin:0;color:#3B82F6">📘 WordPress</div></div><div class="bc">${{hd.replace(b.colLabel,'Category')}}${{mkRows(wpRows,wMax)}}${{mkRows(otherRows,wMax)}}<div class="tr ${{mc}}"><div class="bl">SUBTOTAL</div><div></div><div class="bv">${{fmt(wTotal+oTotal)}}</div><div class="bv sec">${{fmt(wpRows.concat(otherRows).reduce((s,r)=>s+r.websites,0))}}</div>${{hM?`<div class="bv mkt">${{fmt(wTotMkt)}}</div><div class="bv new">${{fmt(wTotNew)}}</div>`:''}}</div></div></div>
</div>`;
}}}}}}
let pnl='';for(const[pk,p] of Object.entries(P)){{let mH='';p.months.forEach((m,mi)=>{{const op=oM[pk+'-'+mi];mH+=`<div class="mr"><div class="mh" onclick="tM('${{pk}}',${{mi}})"><div class="ml"><span class="ma ${{op?'op':''}}">&#9654;</span><span class="mm">${{m.month}}</span><span class="me">${{m.entries.length}} entries</span></div><span class="mc">${{fmt(m.total)}}</span></div><div class="md ${{op?'op':''}}">${{m.entries.map(e=>`<div class="dr"><div class="dd">${{e.date}}</div><div><span class="dt" style="background:${{cc(e.category)}}15;color:${{cc(e.category)}};border:1px solid ${{cc(e.category)}}40">${{e.category}}</span></div><div class="dc">${{fmt(e.count)}}</div></div>`).join('')}}<div class="msw"><span>Subtotal</span><span>${{fmt(m.total)}}</span></div></div></div>`}});pnl+=`<div class="pp ${{aP===pk?'vis':''}}"><div class="pph"><span class="ppn">${{p.name}}</span></div>${{mH}}</div>`}}
const pks=Object.keys(P);const pbt=pks.map(k=>`<button class="tb ${{aP===k?'on':''}}" onclick="sP('${{k}}')">${{P[k].name}}</button>`).join('');
document.getElementById('app').innerHTML=`
<div class="hdr"><div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap"><img src="${{aB==='tcs'?'https://www.google.com/s2/favicons?domain=thecommerceshop.com&sz=64':aB==='drupal'?'https://www.google.com/s2/favicons?domain=thebinaryworks.com&sz=64':aB==='conversionbox'?'https://www.google.com/s2/favicons?domain=conversionbox.ai&sz=64':''}}" style="width:32px;height:32px;border-radius:8px${{isO?';display:none':''}}" /><h1>${{isO?'Others':b?.name||''}} <span class="hl">Lead Database</span></h1><button onclick="window.open(window.location.href.split('?')[0]+'?refresh=1','_self')" style="padding:4px 12px;background:#10B981;color:#fff;border:none;border-radius:6px;font-size:10px;font-weight:700;font-family:inherit;cursor:pointer;white-space:nowrap">🔄 Refresh</button></div>
<div class="tg"><button class="tb ${{aB==='tcs'?'on':''}}" onclick="sB('tcs')"><img src="https://www.google.com/s2/favicons?domain=thecommerceshop.com&sz=16" style="width:14px;height:14px;vertical-align:middle;margin-right:4px;border-radius:2px"/>TCS</button><button class="tb ${{aB==='drupal'?'on':''}}" onclick="sB('drupal')"><img src="https://www.google.com/s2/favicons?domain=thebinaryworks.com&sz=16" style="width:14px;height:14px;vertical-align:middle;margin-right:4px;border-radius:2px"/>BinaryWorks</button><button class="tb ${{aB==='conversionbox'?'on':''}}" onclick="sB('conversionbox')"><img src="https://www.google.com/s2/favicons?domain=conversionbox.ai&sz=16" style="width:14px;height:14px;vertical-align:middle;margin-right:4px;border-radius:2px"/>ConversionBox</button><button class="tb ${{aB==='others'?'on':''}}" onclick="sB('others')">Others</button></div></div>
${{main}}
${{pks.length?`<div class="cw"><button class="cb ${{cO?'on':''}}" onclick="tC()"><span class="ar">&#9660;</span>Individual Contribution</button></div>
<div class="cs ${{cO?'op':''}}"><div class="pw"><div class="tg">${{pbt}}</div></div>${{aP&&P[aP]?chart(P[aP],t):''}}
${{(()=>{{const sd=SL[aP];if(!sd)return '';const mos=Object.keys(sd).sort((a,b)=>{{const mo={{'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}};return (mo[a.toLowerCase().slice(0,3)]||0)-(mo[b.toLowerCase().slice(0,3)]||0)}});const total=Object.values(sd).reduce((s,v)=>s+v,0);const smx=Math.max(...Object.values(sd));const sst=4;const ssv=Math.ceil(smx/sst/1000)*1000;const scl=ssv*sst;
let sg='',sy='';for(let i=0;i<=sst;i++){{const pc=(i/sst)*100,v=ssv*i;sg+=`<div class="gl" style="bottom:${{pc}}%"></div>`;sy+=`<div class="yl">${{v>=1000?(v/1000).toFixed(0)+'K':v}}</div>`}}
const sbs=mos.map(m=>{{const h=scl>0?(sd[m]/scl)*100:0;return`<div class="gbg"><div class="gb" style="height:${{Math.max(h,3)}}%;background:linear-gradient(180deg,#FBBF24,#F59E0B);box-shadow:0 0 8px rgba(245,158,11,0.2)"><div class="gv">${{fmt(sd[m])}}</div><div class="gbl">${{m}}</div></div></div>`}}).join('');
return `<div class="gp"><div class="gt" style="color:#F59E0B">📱 Mobile Numbers Enriched & Shared to Sales: ${{fmt(total)}}</div><div class="ga"><div class="gy">${{sy}}</div><div class="gg">${{sg}}</div><div class="gbs">${{sbs}}</div></div></div>`;}})()??(()=>'')()}}
${{pnl}}
<div style="text-align:center;padding:16px 0"><button onclick="window.location.reload()" style="padding:6px 16px;background:#F59E0B;color:#fff;border:none;border-radius:6px;font-size:11px;font-weight:700;font-family:inherit;cursor:pointer">🔄 Refresh Sheet Data</button></div>
</div>`:''}}
<div class="ft">CommerceShop Lead Database Dashboard</div>`}}
render();
</script></body></html>"""


# ════════════════════════════════════════
# MAIN
# ════════════════════════════════════════

# Initialize session state
if "hub_data" not in st.session_state:
    st.session_state.hub_data = None
if "data_source" not in st.session_state:
    st.session_state.data_source = "csv"

# Check for refresh trigger via URL parameter
if st.query_params.get("refresh") == "1":
    if HUBSPOT_SERVICE_KEY:
        with st.spinner("Pulling live data from HubSpot..."):
            st.session_state.hub_data = fetch_hubspot_counts()
            st.session_state.data_source = "hubspot"
        st.query_params.clear()
        st.rerun()
        if not HUBSPOT_SERVICE_KEY:
            st.error("Add HUBSPOT_API_KEY in Streamlit Settings > Secrets first.")
        else:
            with st.spinner("Pulling live data from HubSpot (30 to 60 seconds)..."):
                st.session_state.hub_data = fetch_hubspot_counts()
                st.session_state.data_source = "hubspot"
            st.rerun()

# Load data
data = load_all()

# If HubSpot refresh was clicked, override brand data
if st.session_state.hub_data and st.session_state.data_source == "hubspot":
    hub_live = st.session_state.hub_data
    for b in ["tcs", "drupal", "conversionbox"]:
        data["brands"][b]["rows"] = hub_live[b]["rows"]
        data["brands"][b]["totals"] = hub_live[b]["totals"]
    data["products"] = hub_live.get("products", data["products"])
    data["others"] = hub_live["others"]
    data["overlap"] = hub_live["overlap"]
    # Email marketing cross-reference not available via API
    data["email_mkt"] = {"tcs": None, "drupal": None, "conversionbox": None}

html = build_html(data)
components.html(html, height=1800, scrolling=True)
